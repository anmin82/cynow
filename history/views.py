from datetime import datetime, timedelta, date
import json
import re

from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.utils import timezone

from history.models import (
    HistInventorySnapshot,
    HistSnapshotRequest,
    SnapshotRequestStatus,
    SnapshotType,
)
from django.db import connection
from core.repositories.history_repository import HistoryRepository
from core.repositories.view_repository import ViewRepository


def history(request):
    """상태 변경 이력/집계 (tr_cylinder_status_histories 기반)"""
    start_date, end_date = _get_date_range(request, default_days=30)

    # 필터 파라미터
    filters = {
        "cylinder_no": request.GET.get("cylinder_no", "").strip(),
        "move_code": request.GET.get("move_code", "").strip(),
        "position_user_name": request.GET.get("position_user_name", "").strip(),
        "move_report_no": request.GET.get("move_report_no", "").strip(),
        "gas_name": request.GET.get("gas_name", "").strip(),
        "cylinder_type_key": request.GET.get("cylinder_type_key", "").strip(),
    }
    # 빈 문자열 제거
    filters = {k: v for k, v in filters.items() if v}

    # 이동코드 분류 (실제 존재 코드에 한정)
    move_code_sets = HistoryRepository.get_move_code_sets()
    available_move_codes = HistoryRepository.get_available_move_codes()
    cylinder_type_options = HistoryRepository.get_cylinder_type_options()
    move_code_names = HistoryRepository.get_move_code_options(sorted(available_move_codes))

    if not filters.get("cylinder_type_key"):
        return render(
            request,
            "history/history.html",
            {
                "histories": [],
                "start_date": start_date,
                "end_date": end_date,
                "filters": filters,
                "available_move_codes": sorted(available_move_codes),
                "move_code_names": move_code_names,
                "cylinder_type_options": cylinder_type_options,
                "weekly_total": {},
                "monthly_total": {},
                "weekly_chart": json.dumps({"labels": [], "inbound": [], "ship": [], "charge": []}),
                "monthly_chart": json.dumps({"labels": [], "inbound": [], "ship": [], "charge": []}),
                "error_message": "대시보드의 용기종류를 선택해서 조회해주세요.",
            },
        )

    # 이력 조회
    histories = HistoryRepository.fetch_history(
        start_date=start_date,
        end_date=end_date,
        filters=filters,
        limit=500,
        offset=0,
    )

    # 집계 (주/월)
    cylinder_type_key = filters.get("cylinder_type_key")
    weekly_summary = HistoryRepository.get_period_summary(
        period="week",
        start_date=start_date,
        end_date=end_date,
        code_sets=move_code_sets,
        cylinder_type_key=cylinder_type_key,
    )
    monthly_summary = HistoryRepository.get_period_summary(
        period="month",
        start_date=start_date,
        end_date=end_date,
        code_sets=move_code_sets,
        cylinder_type_key=cylinder_type_key,
    )

    def _aggregate_latest(summary_rows):
        if not summary_rows:
            return {}
        latest_bucket = summary_rows[0]["bucket"]
        totals = {"ship_cnt": 0, "inbound_cnt": 0, "charge_cnt": 0, "maint_out_cnt": 0, "maint_in_cnt": 0}
        for row in summary_rows:
            if row["bucket"] != latest_bucket:
                continue
            for k in totals.keys():
                totals[k] += row.get(k, 0) or 0
        totals["bucket"] = latest_bucket
        return totals

    weekly_total = _aggregate_latest(weekly_summary)
    monthly_total = _aggregate_latest(monthly_summary)

    def _chart_data(summary_rows, period: str):
        # 시간순으로 정렬(과거→현재)
        rows_sorted = sorted(summary_rows, key=lambda r: r["bucket"])
        labels = []
        for r in rows_sorted:
            b = r.get("bucket")
            if not b:
                labels.append("")
            else:
                labels.append(b.strftime("%Y-%m-%d") if period == "week" else b.strftime("%Y-%m"))
        inbound = [r.get("inbound_cnt", 0) or 0 for r in rows_sorted]
        ship = [r.get("ship_cnt", 0) or 0 for r in rows_sorted]
        charge = [r.get("charge_cnt", 0) or 0 for r in rows_sorted]
        return {"labels": labels, "inbound": inbound, "ship": ship, "charge": charge}

    weekly_chart = _chart_data(weekly_summary, "week")
    monthly_chart = _chart_data(monthly_summary, "month")

    context = {
        "histories": histories,
        "start_date": start_date,
        "end_date": end_date,
        "filters": filters,
        "available_move_codes": sorted(available_move_codes),
        "move_code_names": move_code_names,
        "cylinder_type_options": cylinder_type_options,
        "weekly_total": weekly_total,
        "monthly_total": monthly_total,
        "weekly_chart": json.dumps(weekly_chart),
        "monthly_chart": json.dumps(monthly_chart),
    }
    return render(request, "history/history.html", context)


def _get_date_range(request, default_days=30):
    """
    날짜 범위는 한번 입력하면(=GET으로 들어오면) 세션에 저장하고,
    reset=1 이 오기 전까지 모든 history 서브메뉴에서 유지한다.
    """
    if request.GET.get("reset") == "1":
        request.session.pop("history_start_date", None)
        request.session.pop("history_end_date", None)

    def _parse_date_param(value: str) -> date:
        """'YYYY-MM-DD' 또는 'YYYY년 M월 D일' 형태를 date로 파싱."""
        if value is None:
            raise ValueError("date param is None")
        v = str(value).strip()
        if not v:
            raise ValueError("empty date param")
        try:
            return datetime.strptime(v, "%Y-%m-%d").date()
        except Exception:
            pass
        m = re.match(r"^\s*(\d{4})\s*년\s*(\d{1,2})\s*월\s*(\d{1,2})\s*일\s*$", v)
        if m:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        raise ValueError(f"Unsupported date format: {v}")

    start_date_param = request.GET.get("start_date")
    end_date_param = request.GET.get("end_date")

    # 1) GET이 있으면 세션 갱신
    if start_date_param and end_date_param:
        start_dt = _parse_date_param(start_date_param)
        end_dt = _parse_date_param(end_date_param)
        # 세션에는 항상 ISO 포맷으로 정규화해서 저장
        request.session["history_start_date"] = start_dt.strftime("%Y-%m-%d")
        request.session["history_end_date"] = end_dt.strftime("%Y-%m-%d")
        return (start_dt, end_dt)

    # 2) 세션이 있으면 사용
    s_start = request.session.get("history_start_date")
    s_end = request.session.get("history_end_date")
    if s_start and s_end:
        try:
            return (
                _parse_date_param(s_start),
                _parse_date_param(s_end),
            )
        except Exception:
            pass

    # 3) 기본값
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=default_days)
    return start_date, end_date


def history_movement(request):
    """용기 입출하 이력 (입하/출하)"""
    start_date, end_date = _get_date_range(request, default_days=30)
    move_code_sets = HistoryRepository.get_move_code_sets()
    available_move_codes = HistoryRepository.get_available_move_codes()
    cylinder_type_options = HistoryRepository.get_cylinder_type_options()
    move_code_names = HistoryRepository.get_move_code_options(sorted(available_move_codes))

    filters = {
        "cylinder_no": request.GET.get("cylinder_no", "").strip(),
        "move_code": request.GET.get("move_code", "").strip(),
        "gas_name": request.GET.get("gas_name", "").strip(),
        "cylinder_type_key": request.GET.get("cylinder_type_key", "").strip(),
    }
    filters = {k: v for k, v in filters.items() if v}

    if not filters.get("cylinder_type_key"):
        return render(
            request,
            "history/movement.html",
            {
                "start_date": start_date,
                "end_date": end_date,
                "filters": filters,
                "available_move_codes": sorted(available_move_codes),
                "move_code_names": move_code_names,
                "cylinder_type_options": cylinder_type_options,
                "histories": [],
                "inbound_types": [],
                "ship_types": [],
                "error_message": "대시보드의 용기종류를 선택해서 조회해주세요.",
            },
        )

    histories = HistoryRepository.fetch_history(
        start_date=start_date,
        end_date=end_date,
        filters=filters,
        limit=500,
        offset=0,
    )

    inbound_codes = move_code_sets.get("inbound", [])
    ship_codes = move_code_sets.get("ship", [])
    inbound_types = HistoryRepository.get_type_counts(inbound_codes, start_date, end_date)
    ship_types = HistoryRepository.get_type_counts(ship_codes, start_date, end_date)

    context = {
        "start_date": start_date,
        "end_date": end_date,
        "filters": filters,
        "available_move_codes": sorted(available_move_codes),
        "move_code_names": move_code_names,
        "cylinder_type_options": cylinder_type_options,
        "histories": histories,
        "inbound_types": inbound_types,
        "ship_types": ship_types,
    }
    return render(request, "history/movement.html", context)


def history_charge(request):
    """충전 이력"""
    start_date, end_date = _get_date_range(request, default_days=60)
    move_code_sets = HistoryRepository.get_move_code_sets()
    available_move_codes = HistoryRepository.get_available_move_codes()
    cylinder_type_options = HistoryRepository.get_cylinder_type_options()
    move_code_names = HistoryRepository.get_move_code_options(sorted(available_move_codes))

    filters = {
        "cylinder_no": request.GET.get("cylinder_no", "").strip(),
        "move_code": request.GET.get("move_code", "").strip(),
        "gas_name": request.GET.get("gas_name", "").strip(),
        "cylinder_type_key": request.GET.get("cylinder_type_key", "").strip(),
    }
    filters = {k: v for k, v in filters.items() if v}

    if not filters.get("cylinder_type_key"):
        return render(
            request,
            "history/charge.html",
            {
                "start_date": start_date,
                "end_date": end_date,
                "filters": filters,
                "available_move_codes": sorted(available_move_codes),
                "move_code_names": move_code_names,
                "cylinder_type_options": cylinder_type_options,
                "histories": [],
                "charge_types": [],
                "error_message": "대시보드의 용기종류를 선택해서 조회해주세요.",
            },
        )

    histories = HistoryRepository.fetch_history(
        start_date=start_date,
        end_date=end_date,
        filters=filters,
        limit=500,
        offset=0,
    )

    charge_codes = move_code_sets.get("charge", [])
    charge_types = HistoryRepository.get_type_counts(charge_codes, start_date, end_date)

    context = {
        "start_date": start_date,
        "end_date": end_date,
        "filters": filters,
        "available_move_codes": sorted(available_move_codes),
        "move_code_names": move_code_names,
        "cylinder_type_options": cylinder_type_options,
        "histories": histories,
        "charge_types": charge_types,
    }
    return render(request, "history/charge.html", context)


def history_clf3(request):
    """CLF3 출하/충전 누적 확인"""
    move_code_sets = HistoryRepository.get_move_code_sets()
    ship_codes = move_code_sets.get("ship", [])
    charge_codes = move_code_sets.get("charge", [])

    # CLF3 출하 누적(전체 누적)
    clf3_ship = HistoryRepository.get_clf3_ship_counts(
        ship_codes=ship_codes,
    )

    # CLF3 충전 횟수 (전체 누적)
    charge_counts = []
    if charge_codes:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT 
                    h."CYLINDER_NO" AS cylinder_no,
                    COUNT(*) AS charge_cnt
                FROM "fcms_cdc"."tr_cylinder_status_histories" h
                LEFT JOIN cy_cylinder_current c
                    ON RTRIM(h."CYLINDER_NO") = RTRIM(c.cylinder_no)
                WHERE h."MOVE_CODE" = ANY(%s)
                  AND c.dashboard_gas_name ILIKE 'CLF3%%'
                GROUP BY h."CYLINDER_NO"
                ORDER BY charge_cnt DESC
                """,
                [
                    charge_codes,
                ],
            )
            rows = cursor.fetchall()
            charge_counts = [{"cylinder_no": r[0], "charge_cnt": r[1]} for r in rows]

    context = {
        "clf3_ship": clf3_ship,
        "charge_counts": charge_counts,
    }
    return render(request, "history/clf3.html", context)


def history_trend(request):
    """용기 수요 추이 (입하/출하/충전 그래프 전용)"""
    start_date, end_date = _get_date_range(request, default_days=90)
    move_code_sets = HistoryRepository.get_move_code_sets()
    cylinder_type_options = HistoryRepository.get_cylinder_type_options()
    cylinder_type_key = request.GET.get("cylinder_type_key", "").strip()

    if not cylinder_type_key:
        return render(
            request,
            "history/trend.html",
            {
                "start_date": start_date,
                "end_date": end_date,
                "cylinder_type_key": "",
                "cylinder_type_options": cylinder_type_options,
                "weekly_chart": json.dumps({"labels": [], "inbound": [], "ship": [], "charge": []}),
                "monthly_chart": json.dumps({"labels": [], "inbound": [], "ship": [], "charge": []}),
                "monthly_ship_stock_chart": json.dumps({"labels": [], "ship_stock": []}),
                "monthly_occupancy_chart": json.dumps({"labels": [], "available": [], "process": [], "product": [], "ship": [], "unavailable": [], "unknown": []}),
                "occupancy_source": "",
                "yearly_chart": json.dumps({"labels": [], "inbound": [], "ship": [], "charge": []}),
                "error_message": "대시보드의 용기종류를 선택해서 조회해주세요.",
            },
        )

    weekly_summary = HistoryRepository.get_period_summary(
        period="week",
        start_date=start_date,
        end_date=end_date,
        code_sets=move_code_sets,
        cylinder_type_key=cylinder_type_key,
    )
    monthly_summary = HistoryRepository.get_period_summary(
        period="month",
        start_date=start_date,
        end_date=end_date,
        code_sets=move_code_sets,
        cylinder_type_key=cylinder_type_key,
    )

    # 월말(해당 월 마지막 스냅샷) 기준 "출하중(재고)" 총량 추이
    # 표준 상태명은 '출하'지만, 환경에 따라 '출하중'이 섞일 수 있어 둘 다 허용
    ship_stock_rows = HistoryRepository.get_month_end_status_qty(
        cylinder_type_key=cylinder_type_key,
        statuses=["출하", "출하중"],
        start_date=start_date,
        end_date=end_date,
        snapshot_type="DAILY",
    )

    # 월간 점유율(%) 추이: 월별 마지막 스냅샷 기준 상태 그룹 총량
    occupancy_source = "snapshot"
    occupancy_rows = HistoryRepository.get_period_end_occupancy_summary(
        period="month",
        cylinder_type_key=cylinder_type_key,
        start_date=start_date,
        end_date=end_date,
        snapshot_type="DAILY",
    )
    # 스냅샷이 없으면 히스토리 기반으로 월말 상태를 '추정'하여 점유율을 만든다 (2025 전체 대략 목적)
    if not occupancy_rows:
        occupancy_rows = HistoryRepository.get_month_end_occupancy_from_histories(
            cylinder_type_key=cylinder_type_key,
            start_date=start_date,
            end_date=end_date,
        )
        occupancy_source = "history"

    def _format_label(period: str, bucket):
        if not bucket:
            return ""
        if period == "week":
            iso_week = bucket.isocalendar().week
            return f"{bucket.strftime('%Y')}-W{iso_week:02d}"
        if period == "month":
            return bucket.strftime("%Y-%m")
        return bucket.strftime("%Y-%m-%d")

    for r in weekly_summary:
        r["label"] = _format_label("week", r.get("bucket"))
    for r in monthly_summary:
        r["label"] = _format_label("month", r.get("bucket"))

    def _chart_data(summary_rows, period: str):
        rows_sorted = sorted(summary_rows, key=lambda r: r["bucket"])
        labels = [r.get("label") or _format_label(period, r.get("bucket")) for r in rows_sorted]
        inbound = [r.get("inbound_cnt", 0) or 0 for r in rows_sorted]
        ship = [r.get("ship_cnt", 0) or 0 for r in rows_sorted]
        charge = [r.get("charge_cnt", 0) or 0 for r in rows_sorted]
        return {"labels": labels, "inbound": inbound, "ship": ship, "charge": charge}

    weekly_chart = _chart_data(weekly_summary, "week")
    monthly_chart = _chart_data(monthly_summary, "month")
    monthly_ship_stock_chart = {
        "labels": [_format_label("month", r.get("bucket")) for r in ship_stock_rows],
        "ship_stock": [r.get("qty", 0) or 0 for r in ship_stock_rows],
    }
    # 점유율(%)로 변환
    monthly_occupancy_chart = {
        "labels": [_format_label("month", r.get("bucket")) for r in occupancy_rows],
        "available": [],
        "process": [],
        "product": [],
        "ship": [],
        "unavailable": [],
        "unknown": [],
    }
    for r in occupancy_rows:
        total = (r.get("total_qty", 0) or 0) or 0
        def pct(v):
            return round((float(v or 0) / total * 100.0), 2) if total > 0 else 0
        monthly_occupancy_chart["available"].append(pct(r.get("available_qty")))
        monthly_occupancy_chart["process"].append(pct(r.get("process_qty")))
        monthly_occupancy_chart["product"].append(pct(r.get("product_qty")))
        monthly_occupancy_chart["ship"].append(pct(r.get("ship_qty")))
        monthly_occupancy_chart["unavailable"].append(pct(r.get("unavailable_qty")))
        monthly_occupancy_chart["unknown"].append(pct(r.get("unknown_qty")))

    # 출하 상태 병수 라인차트도 스냅샷이 비어있으면(=hist_snapshot 0) 점유 집계의 ship_qty를 활용
    if (not ship_stock_rows) and occupancy_rows:
        monthly_ship_stock_chart = {
            "labels": [_format_label("month", r.get("bucket")) for r in occupancy_rows],
            "ship_stock": [r.get("ship_qty", 0) or 0 for r in occupancy_rows],
        }
    def _totals(rows):
        return {
            "inbound": sum(r.get("inbound_cnt", 0) or 0 for r in rows),
            "ship": sum(r.get("ship_cnt", 0) or 0 for r in rows),
            "charge": sum(r.get("charge_cnt", 0) or 0 for r in rows),
        }

    weekly_total_row = _totals(weekly_summary)
    monthly_total_row = _totals(monthly_summary)

    context = {
        "start_date": start_date,
        "end_date": end_date,
        "cylinder_type_key": cylinder_type_key,
        "cylinder_type_options": cylinder_type_options,
        "weekly_summary": weekly_summary,
        "monthly_summary": monthly_summary,
        "weekly_total_row": weekly_total_row,
        "monthly_total_row": monthly_total_row,
        "weekly_chart": json.dumps(weekly_chart),
        "monthly_chart": json.dumps(monthly_chart),
        "monthly_ship_stock_chart": json.dumps(monthly_ship_stock_chart),
        "monthly_occupancy_chart": json.dumps(monthly_occupancy_chart),
        "occupancy_source": occupancy_source,
    }
    return render(request, "history/trend.html", context)


def history_trend_export(request):
    """수요 추이(입하/출하/충전) 엑셀 다운로드"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    start_date, end_date = _get_date_range(request, default_days=90)
    move_code_sets = HistoryRepository.get_move_code_sets()
    cylinder_type_key = request.GET.get("cylinder_type_key", "").strip()

    if not cylinder_type_key:
        messages.error(request, "용기종류를 선택해주세요.")
        return redirect("history:history_trend")

    weekly_summary = HistoryRepository.get_period_summary(
        period="week",
        start_date=start_date,
        end_date=end_date,
        code_sets=move_code_sets,
        cylinder_type_key=cylinder_type_key,
    )
    monthly_summary = HistoryRepository.get_period_summary(
        period="month",
        start_date=start_date,
        end_date=end_date,
        code_sets=move_code_sets,
        cylinder_type_key=cylinder_type_key,
    )
    def _format_label(period: str, bucket):
        if not bucket:
            return ""
        if period == "week":
            iso_week = bucket.isocalendar().week
            return f"{bucket.strftime('%Y')}-W{iso_week:02d}"
        if period == "month":
            return bucket.strftime("%Y-%m")
        return bucket.strftime("%Y-%m-%d")

    def write_sheet(ws, title, rows, period):
        ws.title = title
        headers = ["기간", "입하", "출하", "충전"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        for r in rows:
            ws.append(
                [
                    _format_label(period, r.get("bucket")),
                    r.get("inbound_cnt", 0) or 0,
                    r.get("ship_cnt", 0) or 0,
                    r.get("charge_cnt", 0) or 0,
                ]
            )

    wb = Workbook()
    ws_week = wb.active
    write_sheet(ws_week, "주간", weekly_summary, "week")
    write_sheet(wb.create_sheet(), "월간", monthly_summary, "month")

    filename = f"cynow_trend_{start_date}_{end_date}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
@permission_required('cynow.can_edit_plan', raise_exception=True)
def manual_snapshot(request):
    """수동 스냅샷 저장"""
    # 최근 스냅샷 요청 기록
    snapshot_requests = HistSnapshotRequest.objects.filter(
        requested_by=request.user
    ).order_by('-requested_at')[:10]
    
    if request.method == 'POST':
        snapshot_datetime = timezone.now()
        
        try:
            # VIEW에서 현재 인벤토리 데이터 조회
            inventory_data = ViewRepository.get_inventory_view()
            
            inserted_count = 0
            skipped_count = 0
            
            for row in inventory_data:
                try:
                    # source_view_updated_at을 timezone aware로 변환
                    source_updated_at = row.get('updated_at')
                    if source_updated_at and isinstance(source_updated_at, str):
                        try:
                            from datetime import datetime
                            dt = datetime.strptime(source_updated_at, '%Y-%m-%d %H:%M:%S')
                            source_updated_at = timezone.make_aware(dt)
                        except:
                            source_updated_at = None
                    elif source_updated_at and not timezone.is_aware(source_updated_at):
                        source_updated_at = timezone.make_aware(source_updated_at)
                    
                    HistInventorySnapshot.objects.create(
                        snapshot_datetime=snapshot_datetime,
                        snapshot_type=SnapshotType.MANUAL,
                        cylinder_type_key=row.get('cylinder_type_key', ''),
                        gas_name=row.get('gas_name', ''),
                        capacity=row.get('capacity'),
                        valve_spec=row.get('valve_spec'),
                        cylinder_spec=row.get('cylinder_spec'),
                        usage_place=row.get('usage_place'),
                        status=row.get('status', ''),
                        location=row.get('location', ''),
                        qty=row.get('qty', 0),
                        source_view_updated_at=source_updated_at,
                        created_by=request.user,
                    )
                    inserted_count += 1
                except Exception as e:
                    skipped_count += 1
            
            # 성공 기록
            HistSnapshotRequest.objects.create(
                requested_at=snapshot_datetime,
                requested_by=request.user,
                reason=request.POST.get('reason', '수동 스냅샷'),
                status=SnapshotRequestStatus.SUCCESS,
                message=f'{inserted_count} records inserted, {skipped_count} skipped'
            )
            
            messages.success(request, f'스냅샷이 저장되었습니다. ({inserted_count}건)')
            return redirect('history:history')
        except Exception as e:
            # 실패 기록
            HistSnapshotRequest.objects.create(
                requested_at=snapshot_datetime,
                requested_by=request.user,
                reason=request.POST.get('reason', '수동 스냅샷'),
                status=SnapshotRequestStatus.FAILED,
                message=str(e)
            )
            messages.error(request, f'스냅샷 저장 실패: {e}')
    
    context = {
        'snapshot_requests': snapshot_requests,
    }
    return render(request, 'history/manual_snapshot.html', context)


def export_excel(request):
    """상태 이력 엑셀 다운로드"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    # 기간
    start_date_param = request.GET.get("start_date")
    end_date_param = request.GET.get("end_date")
    if not start_date_param or not end_date_param:
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=30)
    else:
        # _get_date_range와 동일하게 한국어 날짜 포맷도 허용
        def _parse_date_param(value: str) -> date:
            v = str(value).strip()
            try:
                return datetime.strptime(v, "%Y-%m-%d").date()
            except Exception:
                pass
            m = re.match(r"^\s*(\d{4})\s*년\s*(\d{1,2})\s*월\s*(\d{1,2})\s*일\s*$", v)
            if m:
                return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            raise ValueError(f"Unsupported date format: {v}")

        start_date = _parse_date_param(start_date_param)
        end_date = _parse_date_param(end_date_param)

    # 필터
    filters = {
        "cylinder_no": request.GET.get("cylinder_no", "").strip(),
        "move_code": request.GET.get("move_code", "").strip(),
        "condition_code": request.GET.get("condition_code", "").strip(),
        "program_id": request.GET.get("program_id", "").strip(),
        "location_code": request.GET.get("location_code", "").strip(),
        "position_user_name": request.GET.get("position_user_name", "").strip(),
        "move_report_no": request.GET.get("move_report_no", "").strip(),
        "gas_name": request.GET.get("gas_name", "").strip(),
        "cylinder_type_key": request.GET.get("cylinder_type_key", "").strip(),
    }
    filters = {k: v for k, v in filters.items() if v}

    histories = HistoryRepository.fetch_history(
        start_date=start_date,
        end_date=end_date,
        filters=filters,
        limit=5000,
        offset=0,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "상태이력"

    headers = [
        "이동일시(KST)",
        "용기번호",
        "가스",
        "용량",
        "밸브",
        "스펙",
        "이동코드",
        "상태코드",
        "표준상태",
        "위치",
        "담당자",
        "이동서",
        "제조LOT",
        "충전LOT",
        "Gross(kg)",
        "Net(kg)",
        "Tare(kg)",
        "비고",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for row in histories:
        ws.append(
            [
                row.get("move_date").strftime("%Y-%m-%d %H:%M:%S")
                if row.get("move_date")
                else "",
                row.get("cylinder_no") or "",
                row.get("gas_name") or "",
                row.get("capacity") or "",
                row.get("valve_spec") or "",
                row.get("cylinder_spec") or "",
                row.get("move_code") or "",
                row.get("condition_code") or "",
                row.get("standard_status") or "",
                row.get("location_code") or "",
                row.get("move_staff_name") or "",
                row.get("move_report_no") or "",
                row.get("manufacture_lot") or "",
                row.get("filling_lot") or "",
                row.get("gross_weight") or "",
                row.get("net_weight") or "",
                row.get("tare_weight") or "",
                row.get("remarks") or "",
            ]
        )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response[
        "Content-Disposition"
    ] = f'attachment; filename="cynow_status_history_{start_date}_{end_date}.xlsx"'
    wb.save(response)
    return response
