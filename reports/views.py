from django.shortcuts import render
from django.http import HttpResponse
from django.utils import timezone
from django.db import connection
from datetime import timedelta, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from core.repositories.cylinder_repository import CylinderRepository
from history.models import HistInventorySnapshot
import logging

logger = logging.getLogger(__name__)


def weekly_report(request):
    """주간 보고서"""
    # 최근 7일 데이터
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=7)
    
    # 현재 인벤토리
    current_inventory = CylinderRepository.get_inventory_summary()
    
    # 일주일 전 스냅샷
    week_ago_snapshots = HistInventorySnapshot.objects.filter(
        snapshot_datetime__date=start_date,
        snapshot_type='DAILY'
    )
    
    # 현재와 비교를 위한 집계
    current_summary = {}
    week_ago_summary = {}
    
    for row in current_inventory:
        key = f"{row.get('gas_name')}_{row.get('status')}"
        current_summary[key] = row.get('qty', 0)
    
    for snapshot in week_ago_snapshots:
        key = f"{snapshot.gas_name}_{snapshot.status}"
        week_ago_summary[key] = week_ago_summary.get(key, 0) + snapshot.qty
    
    # 비교 데이터 생성
    comparison_data = []
    all_keys = set(current_summary.keys()) | set(week_ago_summary.keys())
    for key in all_keys:
        gas_name, status = key.rsplit('_', 1)
        current_qty = current_summary.get(key, 0)
        week_ago_qty = week_ago_summary.get(key, 0)
        delta = current_qty - week_ago_qty
        comparison_data.append({
            'gas_name': gas_name,
            'status': status,
            'current_qty': current_qty,
            'week_ago_qty': week_ago_qty,
            'delta': delta,
        })
    
    comparison_data.sort(key=lambda x: abs(x['delta']), reverse=True)
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'current_inventory': current_inventory[:50],  # 최대 50개
        'comparison_data': comparison_data[:20],  # Top 20 변동
    }
    return render(request, 'reports/weekly.html', context)


def monthly_report(request):
    """월간 보고서"""
    # 이번 달 데이터
    today = timezone.now().date()
    start_date = today.replace(day=1)  # 이번 달 1일
    end_date = today
    
    # 현재 인벤토리
    current_inventory = CylinderRepository.get_inventory_summary()
    
    # 지난 달 말일 스냅샷
    last_month_end = (start_date - timedelta(days=1))
    last_month_snapshots = HistInventorySnapshot.objects.filter(
        snapshot_datetime__date=last_month_end,
        snapshot_type='DAILY'
    )
    
    # 현재와 비교를 위한 집계
    current_summary = {}
    last_month_summary = {}
    
    for row in current_inventory:
        key = f"{row.get('gas_name')}_{row.get('status')}"
        current_summary[key] = row.get('qty', 0)
    
    for snapshot in last_month_snapshots:
        key = f"{snapshot.gas_name}_{snapshot.status}"
        last_month_summary[key] = last_month_summary.get(key, 0) + snapshot.qty
    
    # 비교 데이터 생성
    comparison_data = []
    all_keys = set(current_summary.keys()) | set(last_month_summary.keys())
    for key in all_keys:
        gas_name, status = key.rsplit('_', 1)
        current_qty = current_summary.get(key, 0)
        last_month_qty = last_month_summary.get(key, 0)
        delta = current_qty - last_month_qty
        comparison_data.append({
            'gas_name': gas_name,
            'status': status,
            'current_qty': current_qty,
            'last_month_qty': last_month_qty,
            'delta': delta,
        })
    
    comparison_data.sort(key=lambda x: abs(x['delta']), reverse=True)
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'current_inventory': current_inventory[:50],  # 최대 50개
        'comparison_data': comparison_data[:20],  # Top 20 변동
    }
    return render(request, 'reports/monthly.html', context)


def export_weekly_excel(request):
    """주간 보고서 Excel 다운로드"""
    # 최근 7일 데이터
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=7)
    
    # 현재 인벤토리
    current_inventory = CylinderRepository.get_inventory_summary()
    
    # 일주일 전 스냅샷
    week_ago_snapshots = HistInventorySnapshot.objects.filter(
        snapshot_datetime__date=start_date,
        snapshot_type='DAILY'
    )
    
    # 현재와 비교를 위한 집계
    current_summary = {}
    week_ago_summary = {}
    
    for row in current_inventory:
        key = f"{row.get('gas_name')}_{row.get('status')}"
        current_summary[key] = row.get('qty', 0)
    
    for snapshot in week_ago_snapshots:
        key = f"{snapshot.gas_name}_{snapshot.status}"
        week_ago_summary[key] = week_ago_summary.get(key, 0) + snapshot.qty
    
    # 비교 데이터 생성
    comparison_data = []
    all_keys = set(current_summary.keys()) | set(week_ago_summary.keys())
    for key in all_keys:
        gas_name, status = key.rsplit('_', 1)
        current_qty = current_summary.get(key, 0)
        week_ago_qty = week_ago_summary.get(key, 0)
        delta = current_qty - week_ago_qty
        comparison_data.append({
            'gas_name': gas_name,
            'status': status,
            'current_qty': current_qty,
            'week_ago_qty': week_ago_qty,
            'delta': delta,
        })
    
    comparison_data.sort(key=lambda x: abs(x['delta']), reverse=True)
    
    # 엑셀 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "주간 보고서"
    
    # 헤더
    headers = ['가스명', '상태', '1주 전', '현재', '변동']
    ws.append(headers)
    
    # 헤더 스타일
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # 데이터
    for item in comparison_data:
        ws.append([
            item['gas_name'],
            item['status'],
            item['week_ago_qty'],
            item['current_qty'],
            item['delta'],
        ])
    
    # 응답 생성
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="cynow_weekly_report_{start_date}_{end_date}.xlsx"'
    wb.save(response)
    
    return response


def export_monthly_excel(request):
    """월간 보고서 Excel 다운로드"""
    # 이번 달 데이터
    today = timezone.now().date()
    start_date = today.replace(day=1)  # 이번 달 1일
    end_date = today
    
    # 현재 인벤토리
    current_inventory = CylinderRepository.get_inventory_summary()
    
    # 지난 달 말일 스냅샷
    last_month_end = (start_date - timedelta(days=1))
    last_month_snapshots = HistInventorySnapshot.objects.filter(
        snapshot_datetime__date=last_month_end,
        snapshot_type='DAILY'
    )
    
    # 현재와 비교를 위한 집계
    current_summary = {}
    last_month_summary = {}
    
    for row in current_inventory:
        key = f"{row.get('gas_name')}_{row.get('status')}"
        current_summary[key] = row.get('qty', 0)
    
    for snapshot in last_month_snapshots:
        key = f"{snapshot.gas_name}_{snapshot.status}"
        last_month_summary[key] = last_month_summary.get(key, 0) + snapshot.qty
    
    # 비교 데이터 생성
    comparison_data = []
    all_keys = set(current_summary.keys()) | set(last_month_summary.keys())
    for key in all_keys:
        gas_name, status = key.rsplit('_', 1)
        current_qty = current_summary.get(key, 0)
        last_month_qty = last_month_summary.get(key, 0)
        delta = current_qty - last_month_qty
        comparison_data.append({
            'gas_name': gas_name,
            'status': status,
            'current_qty': current_qty,
            'last_month_qty': last_month_qty,
            'delta': delta,
        })
    
    comparison_data.sort(key=lambda x: abs(x['delta']), reverse=True)
    
    # 엑셀 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "월간 보고서"
    
    # 헤더
    headers = ['가스명', '상태', '지난 달 말', '현재', '변동']
    ws.append(headers)
    
    # 헤더 스타일
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # 데이터
    for item in comparison_data:
        ws.append([
            item['gas_name'],
            item['status'],
            item['last_month_qty'],
            item['current_qty'],
            item['delta'],
        ])
    
    # 응답 생성
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="cynow_monthly_report_{start_date.strftime("%Y%m")}.xlsx"'
    wb.save(response)
    
    return response


def daily_report(request):
    """일일 보고서 - 오늘 하루 용기 이동 현황"""
    # 날짜 파라미터 (기본: 오늘)
    date_str = request.GET.get('date')
    if date_str:
        try:
            report_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            report_date = timezone.now().date()
    else:
        report_date = timezone.now().date()
    
    # MOVE_CODE 라벨 매핑
    move_code_labels = {
        '00': '신규구매', '01': '신규등록', '10': '입하', '14': '회수완료',
        '16': '회수없음', '17': '재보관', '19': '이상처리', '21': '충전선택', '22': '충전완료',
        '30': '창고출고', '31': '외부충전', '41': '분석중', '42': '분석완료',
        '50': '창고입고', '51': '수주연결', '52': '연결해제', '60': '출하',
        '65': '영업외출하', '70': '반품', '85': '전출', '86': '전입', '99': '폐기',
    }
    
    # 오늘 이동된 용기 조회
    movements = []
    move_summary = {}
    arrival_details = []  # 입하 용기 상세
    arrival_summary = {'total': 0, 'expired': 0, 'expiring_soon': 0}
    
    try:
        with connection.cursor() as cursor:
            # 오늘 이동 내역 조회 (MOVE_DATE는 datetime이므로 DATE 비교) - 용기/제품 마스터 조인
            # 대시보드와 동일하게 ma_items, ma_valve_specs, ma_cylinder_specs 조인
            # EndUser는 cy_cylinder_current의 정책이 적용된 dashboard_enduser 사용
            cursor.execute('''
                SELECT 
                    h."CYLINDER_NO",
                    h."MOVE_CODE",
                    h."MOVE_DATE",
                    h."MOVE_REPORT_NO",
                    h."SUPPLIER_USER_NAME",
                    h."CUSTOMER_USER_NAME",
                    h."POSITION_USER_NAME",
                    h."REMARKS",
                    CONCAT(
                        COALESCE(h."FILLING_LOT_HEADER", ''),
                        COALESCE(h."FILLING_LOT_NO", ''),
                        CASE WHEN h."FILLING_LOT_BRANCH" IS NOT NULL AND h."FILLING_LOT_BRANCH" != '' 
                             THEN '-' || h."FILLING_LOT_BRANCH" 
                             ELSE '' 
                        END
                    ) as filling_lot,
                    COALESCE(i."DISPLAY_NAME", i."FORMAL_NAME", c."ITEM_CODE", '미분류') as gas_name,
                    COALESCE(c."CAPACITY", 0) as capacity,
                    COALESCE(vs."NAME", '') as valve_spec,
                    COALESCE(cs."NAME", '') as cylinder_spec,
                    COALESCE(cc.dashboard_enduser, '') as enduser,
                    c."WITHSTAND_PRESSURE_MAINTE_DATE",
                    c."WITHSTAND_PRESSURE_TEST_TERM"
                FROM fcms_cdc.tr_cylinder_status_histories h
                LEFT JOIN fcms_cdc.ma_cylinders c ON TRIM(h."CYLINDER_NO") = TRIM(c."CYLINDER_NO")
                LEFT JOIN fcms_cdc.ma_items i ON TRIM(c."ITEM_CODE") = TRIM(i."ITEM_CODE")
                LEFT JOIN fcms_cdc.ma_valve_specs vs ON c."VALVE_SPEC_CODE" = vs."VALVE_SPEC_CODE"
                LEFT JOIN fcms_cdc.ma_cylinder_specs cs ON c."CYLINDER_SPEC_CODE" = cs."CYLINDER_SPEC_CODE"
                LEFT JOIN public.cy_cylinder_current cc ON TRIM(h."CYLINDER_NO") = TRIM(cc.cylinder_no)
                WHERE DATE(h."MOVE_DATE") = %s
                ORDER BY h."MOVE_CODE", i."DISPLAY_NAME", h."CYLINDER_NO"
            ''', [report_date])
            
            item_summary = {}  # 제품명별 집계
            move_by_item = {}  # 이동유형+제품명별 집계
            move_detail_stats = {}  # 이동유형별 상세 통계
            today = report_date
            
            # 스펙에서 제거할 문자열들
            remove_patterns = ['general Y', 'HAMAI', 'NERIKI', 'SHOT-Y In-screw']
            
            def clean_spec(spec):
                """스펙에서 불필요한 문자열 제거"""
                if not spec:
                    return ''
                result = spec
                for pattern in remove_patterns:
                    result = result.replace(pattern, '')
                # 연속 공백 정리
                result = ' '.join(result.split())
                return result.strip()
            
            for row in cursor.fetchall():
                move_code = row[1].strip() if row[1] else ''
                move_label = move_code_labels.get(move_code, move_code)
                gas_name = row[9].strip() if row[9] else '미분류'
                capacity = row[10] or 0
                valve_spec = clean_spec(row[11].strip() if row[11] else '')
                cylinder_spec = clean_spec(row[12].strip() if row[12] else '')
                enduser = row[13].strip() if row[13] else ''
                pressure_test_date = row[14]
                pressure_test_term = row[15] or 0
                
                # 제품명: 가스명/용량/밸브/용기/EndUser 형식
                item_name_parts = [gas_name]
                if capacity:
                    item_name_parts.append(f"{int(capacity)}L")
                if valve_spec:
                    item_name_parts.append(valve_spec)
                if cylinder_spec:
                    item_name_parts.append(cylinder_spec)
                if enduser:
                    item_name_parts.append(enduser)
                item_name = ' / '.join(item_name_parts)
                
                # 내압만료 계산
                is_expired = False
                is_expiring_soon = False
                if pressure_test_date and pressure_test_term:
                    expiry_date = pressure_test_date + timedelta(days=pressure_test_term * 365)
                    expiry_date_only = expiry_date.date() if hasattr(expiry_date, 'date') else expiry_date
                    if expiry_date_only < today:
                        is_expired = True
                    elif expiry_date_only < today + timedelta(days=90):
                        is_expiring_soon = True
                
                movements.append({
                    'cylinder_no': row[0].strip() if row[0] else '',
                    'move_code': move_code,
                    'move_label': move_label,
                    'move_date': row[2],
                    'move_report_no': row[3].strip() if row[3] else '',
                    'supplier': row[4].strip() if row[4] else '',
                    'customer': row[5].strip() if row[5] else '',
                    'position': row[6].strip() if row[6] else '',
                    'remarks': row[7].strip() if row[7] else '',
                    'filling_lot': row[8].strip() if row[8] else '',
                    'item_code': gas_name,
                    'capacity': capacity,
                    'item_name': item_name,
                    'is_expired': is_expired,
                    'is_expiring_soon': is_expiring_soon,
                })
                
                # 이동유형별 요약 집계
                if move_label not in move_summary:
                    move_summary[move_label] = {'code': move_code, 'count': 0}
                move_summary[move_label]['count'] += 1
                
                # 제품명별 집계
                if item_name not in item_summary:
                    item_summary[item_name] = {'count': 0, 'capacity': capacity}
                item_summary[item_name]['count'] += 1
                
                # 이동유형+제품명별 집계
                key = f"{move_code}_{item_name}"
                if key not in move_by_item:
                    move_by_item[key] = {'move_code': move_code, 'move_label': move_label, 'item_name': item_name, 'count': 0, 'expired': 0, 'expiring_soon': 0}
                move_by_item[key]['count'] += 1
                if is_expired:
                    move_by_item[key]['expired'] += 1
                if is_expiring_soon:
                    move_by_item[key]['expiring_soon'] += 1
                
                # 이동유형별 상세 통계
                if move_code not in move_detail_stats:
                    move_detail_stats[move_code] = {'label': move_label, 'total': 0, 'expired': 0, 'expiring_soon': 0, 'by_item': {}}
                move_detail_stats[move_code]['total'] += 1
                if is_expired:
                    move_detail_stats[move_code]['expired'] += 1
                if is_expiring_soon:
                    move_detail_stats[move_code]['expiring_soon'] += 1
                if item_name not in move_detail_stats[move_code]['by_item']:
                    move_detail_stats[move_code]['by_item'][item_name] = {'count': 0, 'expired': 0}
                move_detail_stats[move_code]['by_item'][item_name]['count'] += 1
                if is_expired:
                    move_detail_stats[move_code]['by_item'][item_name]['expired'] += 1
            
            # 입하(10) 용기 상세 정보 조회 (내압검사 만료일 포함) - 대시보드와 동일하게 조인
            cursor.execute('''
                SELECT 
                    h."CYLINDER_NO",
                    h."MOVE_DATE",
                    h."SUPPLIER_USER_NAME",
                    h."CUSTOMER_USER_NAME",
                    COALESCE(i."DISPLAY_NAME", i."FORMAL_NAME", c."ITEM_CODE", '미분류') as gas_name,
                    c."CAPACITY",
                    COALESCE(vs."NAME", '') as valve_spec,
                    COALESCE(cs."NAME", '') as cylinder_spec,
                    COALESCE(cc.dashboard_enduser, '') as enduser,
                    c."WITHSTAND_PRESSURE_MAINTE_DATE",
                    c."WITHSTAND_PRESSURE_TEST_TERM",
                    CASE 
                        WHEN c."WITHSTAND_PRESSURE_MAINTE_DATE" IS NULL THEN NULL
                        WHEN c."WITHSTAND_PRESSURE_TEST_TERM" IS NULL THEN NULL
                        ELSE c."WITHSTAND_PRESSURE_MAINTE_DATE" + (c."WITHSTAND_PRESSURE_TEST_TERM" * INTERVAL '1 year')
                    END as pressure_expiry_date
                FROM fcms_cdc.tr_cylinder_status_histories h
                LEFT JOIN fcms_cdc.ma_cylinders c ON TRIM(h."CYLINDER_NO") = TRIM(c."CYLINDER_NO")
                LEFT JOIN fcms_cdc.ma_items i ON TRIM(c."ITEM_CODE") = TRIM(i."ITEM_CODE")
                LEFT JOIN fcms_cdc.ma_valve_specs vs ON c."VALVE_SPEC_CODE" = vs."VALVE_SPEC_CODE"
                LEFT JOIN fcms_cdc.ma_cylinder_specs cs ON c."CYLINDER_SPEC_CODE" = cs."CYLINDER_SPEC_CODE"
                LEFT JOIN public.cy_cylinder_current cc ON TRIM(h."CYLINDER_NO") = TRIM(cc.cylinder_no)
                WHERE DATE(h."MOVE_DATE") = %s
                  AND h."MOVE_CODE" = '10'
                ORDER BY h."CYLINDER_NO"
            ''', [report_date])
            
            today = report_date
            for row in cursor.fetchall():
                # 인덱스: 0=CYLINDER_NO, 1=MOVE_DATE, 2=SUPPLIER, 3=CUSTOMER, 4=gas_name, 
                # 5=CAPACITY, 6=valve_spec, 7=cylinder_spec, 8=enduser,
                # 9=WITHSTAND_PRESSURE_MAINTE_DATE, 10=WITHSTAND_PRESSURE_TEST_TERM, 11=pressure_expiry_date
                expiry_date = row[11]
                is_expired = False
                is_expiring_soon = False
                
                if expiry_date:
                    expiry_date_only = expiry_date.date() if hasattr(expiry_date, 'date') else expiry_date
                    if expiry_date_only < today:
                        is_expired = True
                        arrival_summary['expired'] += 1
                    elif expiry_date_only < today + timedelta(days=90):
                        is_expiring_soon = True
                        arrival_summary['expiring_soon'] += 1
                
                # 제품명: 가스명/용량/밸브/용기/EndUser 형식
                gas_name = row[4].strip() if row[4] else '미분류'
                capacity = row[5] or 0
                valve_spec = clean_spec(row[6].strip() if row[6] else '')
                cylinder_spec = clean_spec(row[7].strip() if row[7] else '')
                enduser = row[8].strip() if row[8] else ''
                
                item_name_parts = [gas_name]
                if capacity:
                    item_name_parts.append(f"{int(capacity)}L")
                if valve_spec:
                    item_name_parts.append(valve_spec)
                if cylinder_spec:
                    item_name_parts.append(cylinder_spec)
                if enduser:
                    item_name_parts.append(enduser)
                item_name = ' / '.join(item_name_parts)
                
                arrival_summary['total'] += 1
                arrival_details.append({
                    'cylinder_no': row[0].strip() if row[0] else '',
                    'move_date': row[1],
                    'supplier': row[2].strip() if row[2] else '',
                    'customer': row[3].strip() if row[3] else '',
                    'item_code': item_name,
                    'capacity': capacity,
                    'pressure_test_date': row[9],
                    'pressure_test_term': row[10] or 0,
                    'pressure_expiry_date': expiry_date,
                    'is_expired': is_expired,
                    'is_expiring_soon': is_expiring_soon,
                })
    
    except Exception as e:
        logger.error(f"일일 보고서 조회 오류: {e}")
    
    # 요약 정렬 (건수 많은 순)
    summary_list = [
        {'label': k, 'code': v['code'], 'count': v['count']}
        for k, v in sorted(move_summary.items(), key=lambda x: -x[1]['count'])
    ]
    
    # 현재 인벤토리
    current_inventory = CylinderRepository.get_inventory_summary()
    
    # 상태별 집계
    status_summary = {}
    total_cylinders = 0
    for row in current_inventory:
        status = row.get('status', '기타')
        qty = row.get('qty', 0)
        if status not in status_summary:
            status_summary[status] = 0
        status_summary[status] += qty
        total_cylinders += qty
    
    # 제품명별 집계 정렬
    item_summary_list = [
        {'item_name': k, 'count': v['count'], 'capacity': v['capacity']}
        for k, v in sorted(item_summary.items(), key=lambda x: -x[1]['count'])
    ]
    
    # 이동유형+제품명별 집계 정렬 (이동코드순, 건수 내림차순)
    move_by_item_list = sorted(
        move_by_item.values(),
        key=lambda x: (x['move_code'], -x['count'])
    )
    
    # 이동유형별 상세 통계 정렬
    move_detail_stats_list = []
    for code, stats in sorted(move_detail_stats.items()):
        by_item_list = [
            {'item_name': k, 'count': v['count'], 'expired': v['expired']}
            for k, v in sorted(stats['by_item'].items(), key=lambda x: -x[1]['count'])
        ]
        move_detail_stats_list.append({
            'code': code,
            'label': stats['label'],
            'total': stats['total'],
            'expired': stats['expired'],
            'expiring_soon': stats['expiring_soon'],
            'by_item': by_item_list,
        })
    
    context = {
        'report_date': report_date,
        'movements': movements,
        'move_summary': summary_list,
        'total_movements': len(movements),
        'status_summary': status_summary,
        'total_cylinders': total_cylinders,
        'arrival_details': arrival_details,
        'arrival_summary': arrival_summary,
        'item_summary': item_summary_list,
        'move_by_item': move_by_item_list,
        'move_detail_stats': move_detail_stats_list,
        'generated_at': timezone.now(),
    }
    return render(request, 'reports/daily.html', context)


def arrival_report(request):
    """입하 보고서 - 오늘 입하된 용기 전문 브리핑용 (A4 PDF 출력)"""
    date_str = request.GET.get('date')
    if date_str:
        try:
            report_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            report_date = timezone.now().date()
    else:
        report_date = timezone.now().date()
    
    # 스펙에서 제거할 문자열들
    remove_patterns = ['general Y', 'HAMAI', 'NERIKI', 'SHOT-Y In-screw']
    
    def clean_spec(spec):
        if not spec:
            return ''
        result = spec
        for pattern in remove_patterns:
            result = result.replace(pattern, '')
        result = ' '.join(result.split())
        return result.strip()
    
    arrivals = []
    summary = {
        'total': 0,
        'expired': 0,
        'expiring_soon': 0,
        'normal': 0,
        'by_item': {},
        'by_enduser': {},
    }
    
    try:
        with connection.cursor() as cursor:
            # 입하(10) 용기 상세 조회
            cursor.execute('''
                SELECT 
                    h."CYLINDER_NO",
                    h."MOVE_DATE",
                    h."SUPPLIER_USER_NAME",
                    h."CUSTOMER_USER_NAME",
                    h."MOVE_REPORT_NO",
                    COALESCE(i."DISPLAY_NAME", i."FORMAL_NAME", c."ITEM_CODE", '미분류') as gas_name,
                    c."CAPACITY",
                    COALESCE(vs."NAME", '') as valve_spec,
                    COALESCE(cs."NAME", '') as cylinder_spec,
                    COALESCE(cc.dashboard_enduser, '') as enduser,
                    c."WITHSTAND_PRESSURE_MAINTE_DATE",
                    c."WITHSTAND_PRESSURE_TEST_TERM",
                    CASE 
                        WHEN c."WITHSTAND_PRESSURE_MAINTE_DATE" IS NULL THEN NULL
                        WHEN c."WITHSTAND_PRESSURE_TEST_TERM" IS NULL THEN NULL
                        ELSE c."WITHSTAND_PRESSURE_MAINTE_DATE" + (c."WITHSTAND_PRESSURE_TEST_TERM" * INTERVAL '1 year')
                    END as pressure_expiry_date,
                    c."WEIGHT" as cylinder_weight
                FROM fcms_cdc.tr_cylinder_status_histories h
                LEFT JOIN fcms_cdc.ma_cylinders c ON TRIM(h."CYLINDER_NO") = TRIM(c."CYLINDER_NO")
                LEFT JOIN fcms_cdc.ma_items i ON TRIM(c."ITEM_CODE") = TRIM(i."ITEM_CODE")
                LEFT JOIN fcms_cdc.ma_valve_specs vs ON c."VALVE_SPEC_CODE" = vs."VALVE_SPEC_CODE"
                LEFT JOIN fcms_cdc.ma_cylinder_specs cs ON c."CYLINDER_SPEC_CODE" = cs."CYLINDER_SPEC_CODE"
                LEFT JOIN public.cy_cylinder_current cc ON TRIM(h."CYLINDER_NO") = TRIM(cc.cylinder_no)
                WHERE DATE(h."MOVE_DATE") = %s
                  AND h."MOVE_CODE" = '10'
                ORDER BY i."DISPLAY_NAME", h."CYLINDER_NO"
            ''', [report_date])
            
            today = report_date
            for row in cursor.fetchall():
                expiry_date = row[12]
                is_expired = False
                is_expiring_soon = False
                pressure_status = '정상'
                
                if expiry_date:
                    expiry_date_only = expiry_date.date() if hasattr(expiry_date, 'date') else expiry_date
                    if expiry_date_only < today:
                        is_expired = True
                        summary['expired'] += 1
                        pressure_status = '내압만료'
                    elif expiry_date_only < today + timedelta(days=90):
                        is_expiring_soon = True
                        summary['expiring_soon'] += 1
                        pressure_status = '만료임박'
                    else:
                        summary['normal'] += 1
                else:
                    summary['normal'] += 1
                
                gas_name = row[5].strip() if row[5] else '미분류'
                capacity = row[6] or 0
                valve_spec = clean_spec(row[7].strip() if row[7] else '')
                cylinder_spec = clean_spec(row[8].strip() if row[8] else '')
                enduser = row[9].strip() if row[9] else ''
                supplier = row[2].strip() if row[2] else ''
                
                # 제품명 조합
                item_name_parts = [gas_name]
                if capacity:
                    item_name_parts.append(f"{int(capacity)}L")
                if valve_spec:
                    item_name_parts.append(valve_spec)
                if cylinder_spec:
                    item_name_parts.append(cylinder_spec)
                if enduser:
                    item_name_parts.append(enduser)
                item_name = ' / '.join(item_name_parts)
                
                # 집계
                summary['total'] += 1
                if item_name not in summary['by_item']:
                    summary['by_item'][item_name] = {'count': 0, 'expired': 0, 'normal': 0}
                summary['by_item'][item_name]['count'] += 1
                if is_expired:
                    summary['by_item'][item_name]['expired'] += 1
                else:
                    summary['by_item'][item_name]['normal'] += 1
                
                if enduser:
                    if enduser not in summary['by_enduser']:
                        summary['by_enduser'][enduser] = 0
                    summary['by_enduser'][enduser] += 1
                
                arrivals.append({
                    'cylinder_no': row[0].strip() if row[0] else '',
                    'move_date': row[1],
                    'supplier': supplier,
                    'customer': row[3].strip() if row[3] else '',
                    'move_report_no': row[4].strip() if row[4] else '',
                    'gas_name': gas_name,
                    'capacity': capacity,
                    'item_name': item_name,
                    'pressure_test_date': row[10],
                    'pressure_expiry_date': expiry_date,
                    'pressure_status': pressure_status,
                    'is_expired': is_expired,
                    'is_expiring_soon': is_expiring_soon,
                    'cylinder_weight': row[13] or 0,
                })
    
    except Exception as e:
        logger.error(f"입하 보고서 조회 오류: {e}")
    
    # 제품별 집계 정렬
    by_item_list = [
        {'item_name': k, 'count': v['count'], 'expired': v['expired'], 'normal': v['normal']}
        for k, v in sorted(summary['by_item'].items(), key=lambda x: -x[1]['count'])
    ]
    
    # 엔드유저별 집계 정렬
    by_enduser_list = [
        {'enduser': k, 'count': v}
        for k, v in sorted(summary['by_enduser'].items(), key=lambda x: -x[1])
    ]
    
    # 현재 보유 가용 용기 현황 (대시보드와 동일)
    from core.utils.view_helper import group_cylinder_types
    current_inventory = CylinderRepository.get_inventory_summary()
    cylinder_types_dict = group_cylinder_types(current_inventory)
    
    # 가용 수량 기준 상위 10개
    inventory_list = sorted(
        [
            {
                'gas_name': v['gas_name'],
                'capacity': v['capacity'],
                'available_qty': v['available_qty'],
                'total_qty': v['total_qty'],
            }
            for v in cylinder_types_dict.values()
        ],
        key=lambda x: -x['available_qty']
    )[:10]
    
    # 전체 가용/총 집계
    total_available = sum(v['available_qty'] for v in cylinder_types_dict.values())
    total_inventory = sum(v['total_qty'] for v in cylinder_types_dict.values())
    
    context = {
        'report_date': report_date,
        'arrivals': arrivals,
        'summary': summary,
        'by_item': by_item_list,
        'by_enduser': by_enduser_list,
        'inventory_list': inventory_list,
        'total_available': total_available,
        'total_inventory': total_inventory,
        'generated_at': timezone.now(),
    }
    return render(request, 'reports/arrival.html', context)
