from django.shortcuts import render, get_object_or_404, redirect
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_POST
from django.contrib import messages
from django.db import connection
from core.repositories.cylinder_repository import CylinderRepository
from core.utils.view_helper import parse_cylinder_spec, parse_valve_spec, parse_usage_place
from core.utils.translation import translate_text
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
from .models import CylinderMemo


_TCS_SHIP_DATE_COL: str | None = None
_TCS_SHIP_DATE_COL_READY: bool = False


def _detect_tcs_ship_date_column(cursor) -> str | None:
    """
    fcms_cdc.tr_latest_cylinder_statuses 에서 출하일자(Shipping Date)로 보이는 컬럼을 탐지한다.
    - 환경/버전별로 컬럼명이 다를 수 있어 런타임에 한 번만 탐지 후 캐시한다.
    """
    global _TCS_SHIP_DATE_COL, _TCS_SHIP_DATE_COL_READY
    if _TCS_SHIP_DATE_COL_READY:
        return _TCS_SHIP_DATE_COL

    try:
        cursor.execute(
            """
            SELECT LOWER(column_name) AS col
            FROM information_schema.columns
            WHERE table_schema = %s
              AND table_name = %s
            """,
            ["fcms_cdc", "tr_latest_cylinder_statuses"],
        )
        cols = [r[0] for r in cursor.fetchall() if r and r[0]]
    except Exception:
        cols = []

    # 우선순위: 흔한 컬럼명들
    priority = [
        "shipping_date",
        "ship_date",
        "shipping_dt",
        "ship_dt",
        "shipped_date",
        "out_date",
        "out_dt",
        "delivery_date",
        "deliver_date",
    ]

    col = None
    for p in priority:
        if p in cols:
            col = p
            break

    # 다음 후보: ship + date 같은 패턴
    if col is None:
        for c in cols:
            if "ship" in c and ("date" in c or "dt" in c):
                col = c
                break

    _TCS_SHIP_DATE_COL = col
    _TCS_SHIP_DATE_COL_READY = True
    return _TCS_SHIP_DATE_COL


def _coerce_to_dateish(value):
    """DB에서 가져온 출하일자를 date/datetime로 최대한 변환 (템플릿 date 필터가 먹도록)."""
    if value is None:
        return None
    if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
        return value  # date/datetime

    s = str(value).strip()
    if not s:
        return None

    # yyyymmdd 같은 숫자형
    if s.isdigit() and len(s) == 8:
        try:
            return datetime.strptime(s, "%Y%m%d").date()
        except Exception:
            pass

    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"):
        try:
            parsed = datetime.strptime(s, fmt)
            return parsed.date() if "H" not in fmt else parsed
        except Exception:
            continue
    return value


def _parse_cylinders_list_request(request):
    """
    리스트/엑셀/QR 출력이 동일한 필터 규칙을 사용하도록 공통 파서로 통합.
    """
    # 검색어 파라미터
    search_query = request.GET.get('search', '').strip()
    
    # 다중 선택 필터 파라미터
    selected_gases = request.GET.getlist('gases')
    selected_locations = request.GET.getlist('locations')
    selected_statuses = request.GET.getlist('statuses')
    
    # 단일 선택(하위 호환)
    gas_name = request.GET.get('gas_name', '')
    status = request.GET.get('status', '')
    location = request.GET.get('location', '')
    valve_spec = request.GET.get('valve_spec', '')
    cylinder_spec = request.GET.get('cylinder_spec', '')
    usage_place = request.GET.get('usage_place', '')
    cylinder_type_key = request.GET.get('cylinder_type_key', '')
    cylinder_type_keys_param = request.GET.get('cylinder_type_keys', '').strip()
    days = request.GET.get('days', '')
    
    sort_by = request.GET.get('sort', 'cylinder_no')
    sort_order = request.GET.get('order', 'asc')
    
    if gas_name and gas_name not in selected_gases:
        selected_gases.append(gas_name)
    if location and location not in selected_locations:
        selected_locations.append(location)
    if status and status not in selected_statuses:
        selected_statuses.append(status)
    
    # 상태 확장/레거시 호환
    if selected_statuses:
        expanded = []
        for s in selected_statuses:
            ss = (s or '').strip()
            if not ss:
                continue
            expanded.append(ss)
            if ss == '보관':
                expanded.append('보관:미회수')
                expanded.append('보관:회수')
            if ss == '분석중':
                expanded.append('분석')
            elif ss == '충전중':
                expanded.append('충전')
            elif ss == '제품':
                # DB/동기화 함수에서 '창입'으로 기록된 케이스도 제품으로 취급
                expanded.append('창입')
            elif ss == '정비대상':
                expanded.append('정비')
        seen = set()
        selected_statuses = [x for x in expanded if not (x in seen or seen.add(x))]

    filters = {}
    if selected_gases:
        if len(selected_gases) == 1:
            filters['gas_name'] = selected_gases[0]
        else:
            filters['gases'] = selected_gases
    if selected_locations:
        if len(selected_locations) == 1:
            filters['location'] = selected_locations[0]
        else:
            filters['locations'] = selected_locations
    if selected_statuses:
        if len(selected_statuses) == 1:
            filters['status'] = selected_statuses[0]
        else:
            filters['statuses'] = selected_statuses
    if valve_spec:
        filters['valve_spec'] = valve_spec
    if cylinder_spec:
        filters['cylinder_spec'] = cylinder_spec
    if usage_place:
        filters['usage_place'] = usage_place
    if cylinder_type_key:
        filters['cylinder_type_key'] = cylinder_type_key
    if cylinder_type_keys_param:
        cylinder_type_keys = [k.strip() for k in cylinder_type_keys_param.split(',') if k.strip()]
        if cylinder_type_keys:
            filters['cylinder_type_keys'] = cylinder_type_keys
    
    days_int = None
    if days:
        try:
            days_int = int(days)
        except Exception:
            days_int = None

    return {
        "search_query": search_query,
        "selected_gases": selected_gases,
        "selected_locations": selected_locations,
        "selected_statuses": selected_statuses,
        "gas_name": gas_name,
        "status": status,
        "location": location,
        "valve_spec": valve_spec,
        "cylinder_spec": cylinder_spec,
        "usage_place": usage_place,
        "cylinder_type_key": cylinder_type_key,
        "cylinder_type_keys_param": cylinder_type_keys_param,
        "days": days,
        "days_int": days_int,
        "sort_by": sort_by,
        "sort_order": sort_order,
        "filters": filters,
    }


def cylinder_list(request):
    """용기번호 리스트"""
    parsed = _parse_cylinders_list_request(request)
    search_query = parsed["search_query"]
    selected_gases = parsed["selected_gases"]
    selected_locations = parsed["selected_locations"]
    selected_statuses = parsed["selected_statuses"]
    gas_name = parsed["gas_name"]
    status = parsed["status"]
    location = parsed["location"]
    valve_spec = parsed["valve_spec"]
    cylinder_spec = parsed["cylinder_spec"]
    usage_place = parsed["usage_place"]
    cylinder_type_key = parsed["cylinder_type_key"]
    cylinder_type_keys_param = parsed["cylinder_type_keys_param"]
    days = parsed["days"]
    days_int = parsed["days_int"]
    sort_by = parsed["sort_by"]
    sort_order = parsed["sort_order"]
    filters = parsed["filters"]
    
    # 페이지네이션
    page = request.GET.get('page', 1)
    try:
        page = int(page)
    except:
        page = 1
    
    per_page = 50
    offset = (page - 1) * per_page
    
    # 전체 개수 조회 (페이지네이션용)
    total_count = CylinderRepository.get_cylinder_count(filters=filters, days=days_int, search_query=search_query)
    
    # 현재 페이지 데이터만 조회 (SQL LIMIT/OFFSET 사용)
    cylinders_list = CylinderRepository.get_cylinder_list(
        filters=filters, 
        limit=per_page, 
        offset=offset,
        days=days_int,
        sort_by=sort_by,
        sort_order=sort_order,
        search_query=search_query
    )

    # 상태 표시 정규화 (레거시 상태를 사용자 표시용 상태로 변환)
    # - DB에는 '분석'이 남아있을 수 있으나 UI에서는 '분석중'으로 보이게 한다.
    for c in cylinders_list:
        try:
            s = (c.get('status', '') or '').strip()
            if s == '분석':
                c['status'] = '분석중'
            elif s == '충전':
                c['status'] = '충전중'
            elif s in ('창입', '倉入', '倉入済'):
                c['status'] = '제품'
        except Exception:
            pass
    
    # 페이지네이션 계산
    import math
    total_pages = math.ceil(total_count / per_page) if total_count > 0 else 1
    
    # 커스텀 페이지네이션 객체 생성
    class CustomPaginator:
        def __init__(self, count, per_page):
            self._count = count
            self.per_page = per_page
        
        @property
        def count(self):
            return self._count
        
        @property
        def num_pages(self):
            return math.ceil(self._count / self.per_page) if self._count > 0 else 1
    
    class CustomPage:
        def __init__(self, object_list, number, paginator, total_count):
            self.object_list = object_list
            self.number = number
            self.paginator = paginator
            self._total_count = total_count
        
        def __iter__(self):
            return iter(self.object_list)
        
        def __len__(self):
            return len(self.object_list)
        
        @property
        def has_previous(self):
            return self.number > 1
        
        @property
        def has_next(self):
            return self.number < self.paginator.num_pages
        
        @property
        def previous_page_number(self):
            return self.number - 1 if self.has_previous else None
        
        @property
        def next_page_number(self):
            return self.number + 1 if self.has_next else None
        
        @property
        def start_index(self):
            if self._total_count == 0:
                return 0
            return (self.number - 1) * self.paginator.per_page + 1
        
        @property
        def end_index(self):
            if self._total_count == 0:
                return 0
            return min(self.number * self.paginator.per_page, self._total_count)
    
    paginator = CustomPaginator(total_count, per_page)
    cylinders = CustomPage(cylinders_list, page, paginator, total_count)
    current_page = page
    page_range = []
    
    if total_pages <= 7:
        # 7페이지 이하면 모두 표시
        page_range = list(range(1, total_pages + 1))
    else:
        if current_page <= 4:
            # 앞쪽: 1 2 3 4 5 ... last
            page_range = list(range(1, 6)) + [None, total_pages]
        elif current_page >= total_pages - 3:
            # 뒤쪽: 1 ... n-4 n-3 n-2 n-1 n
            page_range = [1, None] + list(range(total_pages - 4, total_pages + 1))
        else:
            # 중간: 1 ... cur-1 cur cur+1 ... last
            page_range = [1, None, current_page - 1, current_page, current_page + 1, None, total_pages]
    
    # 필터 옵션 조회 (최적화된 쿼리 사용)
    filter_options = CylinderRepository.get_filter_options()
    
    # 위치 중복 제거 (번역 후 같은 값이 되는 경우)
    unique_locations = []
    seen_translated = set()
    for loc in filter_options['locations']:
        translated = translate_text('location', loc) or loc
        if translated not in seen_translated:
            seen_translated.add(translated)
            unique_locations.append(loc)
    
    # 상태 목록 (UI 표시용, 새 체계 기준)
    # - '보관'은 통합 선택(미회수/회수)을 의미
    # - '분석중'은 레거시 '분석'에 대응
    status_options = ['보관', '충전중', '충전완료', '분석중', '분석완료', '제품', '출하', '이상', '정비대상', '폐기']
    
    # 현재 페이지 용기들 중 메모가 있는 용기 번호 조회
    cylinder_nos = [c.get('cylinder_no') for c in cylinders_list if c.get('cylinder_no')]
    cylinders_with_memo = set()
    if cylinder_nos:
        cylinders_with_memo = set(
            CylinderMemo.objects.filter(
                cylinder_no__in=cylinder_nos,
                is_active=True,
                parent__isnull=True  # 최상위 메모만
            ).values_list('cylinder_no', flat=True)
        )
    
    context = {
        'cylinders': cylinders,
        'total_count': total_count,
        'cylinders_with_memo': cylinders_with_memo,
        'page_range': page_range,
        # 검색어
        'search_query': search_query,
        # 다중 선택 필터
        'selected_gases': selected_gases,
        'selected_locations': selected_locations,
        'selected_statuses': selected_statuses,
        # 단일 선택 (하위 호환)
        'gas_name': gas_name,
        'status': status,
        'location': location,
        'valve_spec': valve_spec,
        'cylinder_spec': cylinder_spec,
        'usage_place': usage_place,
        'cylinder_type_key': cylinder_type_key,
        'cylinder_type_keys': cylinder_type_keys_param,
        'days': days,
        # 정렬
        'sort_by': sort_by,
        'sort_order': sort_order,
        # 필터 옵션
        'gas_names': filter_options['gas_names'],
        'locations': unique_locations,
        'valve_specs': filter_options['valve_specs'],
        'cylinder_specs': filter_options['cylinder_specs'],
        'statuses': status_options,
    }
    return render(request, 'cylinders/list.html', context)


def memoed_cylinders_summary(request):
    """
    메모된 용기 목록 요약 API (모달용)
    - 목적: cylinders 리스트에서 "메모 관리" 모달을 띄우기 위한 데이터 제공
    - 비밀번호/수정/삭제는 개별 상세 페이지에서 처리
    """
    q = (request.GET.get("q") or "").strip()
    limit = request.GET.get("limit", "200")
    try:
        limit_int = max(1, min(int(limit), 500))
    except Exception:
        limit_int = 200

    qs = CylinderMemo.objects.filter(is_active=True, parent__isnull=True)
    if q:
        qs = qs.filter(cylinder_no__icontains=q)

    # 용기별 집계: 메모 수, 최신 작성일, 최신 내용 일부
    from django.db.models import Count, Max

    agg = (
        qs.values("cylinder_no")
        .annotate(memo_count=Count("id"), last_at=Max("created_at"))
        .order_by("-last_at")[:limit_int]
    )

    cylinder_nos = [r["cylinder_no"] for r in agg if r.get("cylinder_no")]
    # 최신 메모 내용 프리뷰 (용기별 1개)
    latest_map = {}
    if cylinder_nos:
        latest_qs = (
            CylinderMemo.objects.filter(is_active=True, parent__isnull=True, cylinder_no__in=cylinder_nos)
            .order_by("cylinder_no", "-created_at")
        )
        # Python에서 그룹별 first
        for m in latest_qs:
            if m.cylinder_no not in latest_map:
                latest_map[m.cylinder_no] = (m.content or "")[:60]

    rows = []
    for r in agg:
        cno = r.get("cylinder_no") or ""
        rows.append(
            {
                "cylinder_no": cno,
                "memo_count": int(r.get("memo_count") or 0),
                "last_at": (r.get("last_at").strftime("%Y-%m-%d %H:%M") if r.get("last_at") else ""),
                "preview": latest_map.get(cno, ""),
                "detail_url": f"/cynow/cylinders/{cno}/",
            }
        )

    return JsonResponse({"success": True, "count": len(rows), "rows": rows})


def cylinder_detail(request, cylinder_no):
    """용기 상세보기"""
    # URL에서 받은 용기번호의 앞뒤 공백 제거
    cylinder_no = cylinder_no.strip()
    
    filters = {'cylinder_no': cylinder_no}
    cylinders = CylinderRepository.get_cylinder_list(filters=filters, limit=1)
    
    if not cylinders:
        from django.http import Http404
        raise Http404("용기를 찾을 수 없습니다.")
    
    cylinder = cylinders[0]

    # FCMS 출하일자: tr_latest_cylinder_statuses의 shipping date 컬럼을 그대로 사용 (가장 최신/정합)
    try:
        ship_dt = None
        with connection.cursor() as cursor:
            col = _detect_tcs_ship_date_column(cursor)
            if col:
                # 식별자 안전성 체크(동적 컬럼 사용)
                import re

                if re.match(r"^[a-z0-9_]+$", col):
                    cursor.execute(
                        f"""
                        SELECT tcs."{col.upper()}" AS ship_date
                        FROM "fcms_cdc"."tr_latest_cylinder_statuses" tcs
                        WHERE RTRIM(tcs."CYLINDER_NO") = RTRIM(%s)
                        LIMIT 1
                        """,
                        [cylinder_no],
                    )
                    row = cursor.fetchone()
                    ship_dt = row[0] if row else None

        cylinder["ship_date"] = _coerce_to_dateish(ship_dt)
    except Exception:
        cylinder["ship_date"] = None
    
    # 메모 목록 조회 (최상위 메모만, 답글은 prefetch)
    memos = CylinderMemo.objects.filter(
        cylinder_no=cylinder_no,
        parent__isnull=True,
        is_active=True
    ).prefetch_related(
        'replies'
    ).order_by('-created_at')
    
    # 활성화된 답글만 필터링
    for memo in memos:
        memo.active_replies = memo.replies.filter(is_active=True).order_by('created_at')
    
    context = {
        'cylinder': cylinder,
        'memos': memos,
    }
    return render(request, 'cylinders/detail.html', context)


def cylinder_ship_history(request, cylinder_no):
    """
    용기 출하 이력 조회 (모달용 JSON)
    - 항목: 이동서번호(move_report_no), 출하일자(move_date)
    """
    cylinder_no = (cylinder_no or "").strip()
    limit = request.GET.get("limit", "50")
    try:
        limit_n = max(1, min(200, int(limit)))
    except Exception:
        limit_n = 50

    try:
        from core.repositories.history_repository import HistoryRepository

        ship_codes = (HistoryRepository.get_move_code_sets() or {}).get("ship") or []
        if not ship_codes:
            # 라벨 기반으로 출하 코드 후보 탐색
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    SELECT DISTINCT TRIM(p."KEY1") AS code
                    FROM "fcms_cdc"."ma_parameters" p
                    WHERE p."KEY1" IS NOT NULL
                      AND (p."KEY2" IS NULL OR TRIM(p."KEY2") = '')
                      AND (p."KEY3" IS NULL OR TRIM(p."KEY3") = '')
                      AND (
                        COALESCE(p."VALUE2", p."VALUE1", p."VALUE3", '') ILIKE %s
                        OR COALESCE(p."VALUE2", p."VALUE1", p."VALUE3", '') ILIKE %s
                        OR COALESCE(p."VALUE2", p."VALUE1", p."VALUE3", '') ILIKE %s
                      )
                    """,
                    ["%출하%", "%出荷%", "%SHIP%"],
                )
                ship_codes = [r[0] for r in cursor.fetchall() if r and r[0]]
        if not ship_codes:
            ship_codes = ["60"]

        # 0-padding 제거 비교용 코드 집합
        ship_codes_norm = []
        for c in ship_codes:
            s = str(c).strip() if c is not None else ""
            if not s:
                continue
            ship_codes_norm.append((s.lstrip("0") or s))
        ship_codes_norm = sorted(set(ship_codes_norm))

        rows = []
        with connection.cursor() as cursor:
            if connection.vendor == "postgresql":
                cursor.execute(
                    """
                    SELECT
                        RTRIM(h."MOVE_REPORT_NO") AS move_report_no,
                        h."MOVE_DATE" AS ship_date
                    FROM "fcms_cdc"."tr_cylinder_status_histories" h
                    WHERE RTRIM(h."CYLINDER_NO") = RTRIM(%s)
                      AND NULLIF(regexp_replace(TRIM(h."MOVE_CODE"::text), '^0+', ''), '') = ANY(%s)
                    ORDER BY h."MOVE_DATE" DESC NULLS LAST, h."HISTORY_SEQ" DESC NULLS LAST
                    LIMIT %s
                    """,
                    [cylinder_no, ship_codes_norm, limit_n],
                )
            else:
                placeholders = ", ".join(["%s"] * len(ship_codes_norm))
                cursor.execute(
                    f"""
                    SELECT
                        TRIM(h."MOVE_REPORT_NO") AS move_report_no,
                        h."MOVE_DATE" AS ship_date
                    FROM "fcms_cdc"."tr_cylinder_status_histories" h
                    WHERE TRIM(h."CYLINDER_NO") = TRIM(%s)
                      AND TRIM(h."MOVE_CODE") IN ({placeholders})
                    ORDER BY h."MOVE_DATE" DESC
                    """,
                    [cylinder_no, *ship_codes_norm],
                )
            for move_report_no, ship_date in cursor.fetchall():
                rows.append(
                    {
                        "move_report_no": (move_report_no or "").strip(),
                        "ship_date": ship_date.isoformat() if hasattr(ship_date, "isoformat") else (str(ship_date) if ship_date else ""),
                    }
                )

        return JsonResponse({"ok": True, "cylinder_no": cylinder_no, "rows": rows})
    except Exception as e:
        return JsonResponse({"ok": False, "cylinder_no": cylinder_no, "rows": [], "error": str(e)})


@require_POST
def memo_create(request, cylinder_no):
    """메모 작성"""
    author_name = request.POST.get('author_name', '').strip()
    password = request.POST.get('password', '').strip()
    content = request.POST.get('content', '').strip()
    
    if not author_name or not content or not password:
        messages.error(request, '작성자명, 암호, 내용을 모두 입력해주세요.')
        return redirect('cylinders:detail', cylinder_no=cylinder_no)
    
    if not password.isdigit() or len(password) != 4:
        messages.error(request, '암호는 4자리 숫자로 입력해주세요.')
        return redirect('cylinders:detail', cylinder_no=cylinder_no)
    
    CylinderMemo.objects.create(
        cylinder_no=cylinder_no,
        author_name=author_name,
        password=password,
        content=content
    )
    
    messages.success(request, '메모가 등록되었습니다.')
    return redirect('cylinders:detail', cylinder_no=cylinder_no)


@require_POST
def memo_reply(request, cylinder_no, memo_id):
    """메모 답글 작성"""
    parent_memo = get_object_or_404(CylinderMemo, id=memo_id, cylinder_no=cylinder_no, is_active=True)
    
    author_name = request.POST.get('author_name', '').strip()
    password = request.POST.get('password', '').strip()
    content = request.POST.get('content', '').strip()
    
    if not author_name or not content or not password:
        messages.error(request, '작성자명, 암호, 내용을 모두 입력해주세요.')
        return redirect('cylinders:detail', cylinder_no=cylinder_no)
    
    if not password.isdigit() or len(password) != 4:
        messages.error(request, '암호는 4자리 숫자로 입력해주세요.')
        return redirect('cylinders:detail', cylinder_no=cylinder_no)
    
    CylinderMemo.objects.create(
        cylinder_no=cylinder_no,
        parent=parent_memo,
        author_name=author_name,
        password=password,
        content=content
    )
    
    messages.success(request, '답글이 등록되었습니다.')
    return redirect('cylinders:detail', cylinder_no=cylinder_no)


@require_POST
def memo_edit(request, cylinder_no, memo_id):
    """메모 수정"""
    memo = get_object_or_404(CylinderMemo, id=memo_id, cylinder_no=cylinder_no, is_active=True)
    
    password = request.POST.get('password', '').strip()
    content = request.POST.get('content', '').strip()
    
    if not password or not content:
        messages.error(request, '암호와 내용을 입력해주세요.')
        return redirect('cylinders:detail', cylinder_no=cylinder_no)
    
    if memo.password != password:
        messages.error(request, '암호가 일치하지 않습니다.')
        return redirect('cylinders:detail', cylinder_no=cylinder_no)
    
    memo.content = content
    memo.save()
    
    messages.success(request, '메모가 수정되었습니다.')
    return redirect('cylinders:detail', cylinder_no=cylinder_no)


@require_POST
def memo_delete(request, cylinder_no, memo_id):
    """메모 삭제 (비활성화)"""
    memo = get_object_or_404(CylinderMemo, id=memo_id, cylinder_no=cylinder_no, is_active=True)
    
    password = request.POST.get('password', '').strip()
    
    if not password:
        messages.error(request, '암호를 입력해주세요.')
        return redirect('cylinders:detail', cylinder_no=cylinder_no)
    
    if memo.password != password:
        messages.error(request, '암호가 일치하지 않습니다.')
        return redirect('cylinders:detail', cylinder_no=cylinder_no)
    
    # 실제 삭제 대신 비활성화
    memo.is_active = False
    memo.save()
    
    # 답글도 함께 비활성화
    memo.replies.update(is_active=False)
    
    messages.success(request, '메모가 삭제되었습니다.')
    return redirect('cylinders:detail', cylinder_no=cylinder_no)


def cylinder_export_excel(request):
    """용기 리스트 엑셀 다운로드"""
    from urllib.parse import quote
    
    parsed = _parse_cylinders_list_request(request)
    search_query = parsed["search_query"]
    filters = parsed["filters"]
    days_int = parsed["days_int"]
    sort_by = parsed["sort_by"]
    sort_order = parsed["sort_order"]
    
    # 전체 데이터 조회 (엑셀용 - 최대 10,000건 제한)
    cylinders = CylinderRepository.get_cylinder_list(
        filters=filters, 
        limit=10000,
        days=days_int,
        sort_by=sort_by,
        sort_order=sort_order,
        search_query=search_query
    )
    
    # 엑셀 워크북 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "용기리스트"
    
    # 스타일 정의
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 헤더 정의
    headers = [
        ('용기번호', 15),
        ('가스명', 20),
        ('밸브규격', 15),
        ('용기규격', 15),
        ('상태', 10),
        ('위치', 20),
        ('제조일', 12),
        ('내압시험일', 12),
        ('검사주기', 10),
        ('내압만료일', 12),
        ('FCMS수정필요', 12),
    ]
    
    # 헤더 작성
    for col, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # 데이터 작성
    for row_num, cylinder in enumerate(cylinders, 2):
        # 밸브규격 파싱
        valve_parsed = parse_valve_spec(cylinder.get('valve_spec', '') or '')
        valve_display = f"{valve_parsed.get('format', '-')}/{valve_parsed.get('material', '-')}"
        
        # 용기규격 파싱
        cylinder_parsed = parse_cylinder_spec(cylinder.get('cylinder_spec', '') or '')
        cylinder_display = f"{cylinder_parsed.get('format', '-')}/{cylinder_parsed.get('material', '-')}"
        
        # 사용처 파싱
        usage_place = parse_usage_place(cylinder.get('usage_place', '') or '')
        
        # 날짜 포맷팅
        manufacture_date = cylinder.get('manufacture_date')
        pressure_test_date = cylinder.get('pressure_test_date')
        pressure_expire_date = cylinder.get('pressure_expire_date')
        
        row_data = [
            cylinder.get('cylinder_no', ''),
            translate_text('gas_name', cylinder.get('gas_name', '')),
            valve_display,
            cylinder_display,
            cylinder.get('status', ''),
            translate_text('location', cylinder.get('location', '')),
            manufacture_date.strftime('%Y-%m-%d') if manufacture_date else '-',
            pressure_test_date.strftime('%Y-%m-%d') if pressure_test_date else '-',
            f"{cylinder.get('pressure_test_term', '-')}년" if cylinder.get('pressure_test_term') else '-',
            pressure_expire_date.strftime('%Y-%m-%d') if pressure_expire_date else '-',
            'FCMS 수정필요' if cylinder.get('needs_fcms_fix') else '',
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            if col == 5:  # 상태 컬럼 가운데 정렬
                cell.alignment = Alignment(horizontal="center")
    
    # 필터 정보 시트 추가
    ws_info = wb.create_sheet(title="필터정보")
    ws_info['A1'] = "다운로드 일시"
    ws_info['B1'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws_info['A2'] = "총 건수"
    ws_info['B2'] = len(cylinders)
    ws_info['A3'] = "적용된 필터"
    
    filter_info = []
    if gas_name:
        filter_info.append(f"가스명: {gas_name}")
    if status:
        filter_info.append(f"상태: {status}")
    if location:
        filter_info.append(f"위치: {location}")
    if valve_spec:
        filter_info.append(f"밸브규격: {valve_spec}")
    if cylinder_spec:
        filter_info.append(f"용기규격: {cylinder_spec}")
    if days:
        filter_info.append(f"기간: 최근 {days}일")
    
    ws_info['B3'] = ", ".join(filter_info) if filter_info else "없음"
    
    # HTTP Response 생성
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # 파일명 생성 (날짜 포함) - RFC 5987 형식으로 인코딩
    filename = f"용기리스트_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filename_ascii = f"cylinders_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filename_encoded = quote(filename)
    response['Content-Disposition'] = f"attachment; filename=\"{filename_ascii}\"; filename*=UTF-8''{filename_encoded}"
    
    wb.save(response)
    return response


def cylinder_export_qr_pdf(request):
    """용기 QR코드 PDF 출력 - A4 용지에 다중열로 출력"""
    from urllib.parse import quote
    from io import BytesIO
    import qrcode
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.utils import ImageReader
    import os
    
    parsed = _parse_cylinders_list_request(request)
    search_query = parsed["search_query"]
    filters = parsed["filters"]
    days_int = parsed["days_int"]
    sort_by = parsed["sort_by"]
    sort_order = parsed["sort_order"]
    
    # 전체 데이터 조회 (최대 1000건 제한 - PDF 크기 고려)
    cylinders = CylinderRepository.get_cylinder_list(
        filters=filters, 
        limit=1000,
        days=days_int,
        sort_by=sort_by,
        sort_order=sort_order,
        search_query=search_query
    )
    
    # A4 사이즈 설정
    page_width, page_height = A4  # 595.27 x 841.89 points
    
    # 레이아웃 설정
    margin_x = 10 * mm  # 좌우 여백
    margin_y = 10 * mm  # 상하 여백
    cols = 10  # 한 줄에 10개
    qr_size = 10 * mm  # QR코드 크기 1cm
    cell_width = (page_width - 2 * margin_x) / cols
    cell_height = 15 * mm  # QR코드 + 텍스트 높이 (약 1.5cm)
    
    # 한 페이지당 행 수 계산
    usable_height = page_height - 2 * margin_y
    rows_per_page = int(usable_height / cell_height)
    items_per_page = cols * rows_per_page
    
    # PDF 생성
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    
    # 한글 폰트 등록 시도
    font_name = 'Helvetica'
    try:
        # Windows 맑은 고딕
        font_path = 'C:/Windows/Fonts/malgun.ttf'
        if os.path.exists(font_path):
            pdfmetrics.registerFont(TTFont('MalgunGothic', font_path))
            font_name = 'MalgunGothic'
    except:
        pass
    
    for idx, cylinder in enumerate(cylinders):
        # 페이지 내 위치 계산
        page_idx = idx % items_per_page
        row = page_idx // cols
        col = page_idx % cols
        
        # 새 페이지 시작
        if idx > 0 and page_idx == 0:
            c.showPage()
        
        # 셀 위치 계산 (왼쪽 위 기준)
        x = margin_x + col * cell_width
        y = page_height - margin_y - (row + 1) * cell_height
        
        cylinder_no = cylinder.get('cylinder_no', '')
        
        # QR코드 생성
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=1,
        )
        qr.add_data(cylinder_no)
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color="black", back_color="white")
        
        # QR코드를 BytesIO에 저장
        qr_buffer = BytesIO()
        qr_img.save(qr_buffer, format='PNG')
        qr_buffer.seek(0)
        
        # QR코드 그리기 (셀 중앙에)
        qr_x = x + (cell_width - qr_size) / 2
        qr_y = y + 4 * mm  # 텍스트 공간 확보
        c.drawImage(ImageReader(qr_buffer), qr_x, qr_y, width=qr_size, height=qr_size)
        
        # 용기번호 텍스트
        c.setFont(font_name, 5)  # 작은 폰트
        text_x = x + cell_width / 2
        text_y = y + 1 * mm
        c.drawCentredString(text_x, text_y, cylinder_no[:12] if len(cylinder_no) > 12 else cylinder_no)
    
    c.save()
    
    # HTTP Response 생성
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    
    # 파일명 생성 (RFC 5987 형식)
    filename = f"QR코드_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    filename_ascii = f"qrcodes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    filename_encoded = quote(filename)
    response['Content-Disposition'] = f"attachment; filename=\"{filename_ascii}\"; filename*=UTF-8''{filename_encoded}"
    
    return response


def smart_search(request):
    """스마트 검색 API - 자연어 검색 및 필터 추천"""
    from .search_parser import parse_natural_query, get_scenario_presets
    import json
    
    query = request.GET.get('q', '').strip()
    
    if not query:
        # 빈 쿼리면 시나리오 프리셋만 반환
        return JsonResponse({
            'success': True,
            'query': '',
            'filters': {},
            'parsed_keywords': [],
            'suggestions': [],
            'scenarios': get_scenario_presets(),
        })
    
    # 자연어 파싱
    parsed = parse_natural_query(query)
    
    # 검색 결과 미리보기 (개수만)
    filters = build_filters_from_parsed(parsed['filters'])
    
    # 전체 개수 조회
    total_count = CylinderRepository.get_cylinder_count(filters=filters)
    
    # 결과 기반 추천 필터 (상위 옵션)
    recommendations = get_search_recommendations(filters)
    
    return JsonResponse({
        'success': True,
        'query': query,
        'filters': parsed['filters'],
        'parsed_keywords': parsed['parsed_keywords'],
        'suggestions': parsed['suggestions'],
        'preview_count': total_count,
        'recommendations': recommendations,
        'scenarios': get_scenario_presets(),
    })


def build_filters_from_parsed(parsed_filters: dict) -> dict:
    """파싱된 필터를 repository 필터 형식으로 변환"""
    filters = {}
    
    if 'statuses' in parsed_filters:
        # 첫 번째 상태만 사용 (다중 상태는 별도 처리 필요)
        if parsed_filters['statuses']:
            filters['status'] = parsed_filters['statuses'][0]
    
    if 'location' in parsed_filters:
        filters['location'] = parsed_filters['location']
    
    if 'gas_keyword' in parsed_filters:
        filters['gas_name'] = parsed_filters['gas_keyword']
    
    return filters


def get_search_recommendations(filters: dict) -> dict:
    """현재 필터 기반 추천 옵션"""
    recommendations = {
        'by_status': {},
        'by_location': {},
    }
    
    try:
        # 상태별 개수
        all_data = CylinderRepository.get_cylinder_list(filters=filters, limit=5000)
        
        status_counts = {}
        location_counts = {}
        
        for item in all_data:
            status = item.get('status', '')
            location = item.get('location', '')
            
            if status:
                status_counts[status] = status_counts.get(status, 0) + 1
            if location:
                location_counts[location] = location_counts.get(location, 0) + 1
        
        # 상위 5개만
        recommendations['by_status'] = dict(sorted(status_counts.items(), key=lambda x: -x[1])[:5])
        recommendations['by_location'] = dict(sorted(location_counts.items(), key=lambda x: -x[1])[:5])
    except Exception:
        pass
    
    return recommendations


def search_autocomplete(request):
    """검색 자동완성 API"""
    from .search_parser import parse_natural_query
    
    query = request.GET.get('q', '').strip()
    
    if len(query) < 2:
        return JsonResponse({'suggestions': []})
    
    # 자연어 파싱
    parsed = parse_natural_query(query)
    
    suggestions = []
    
    # 파싱된 키워드 기반 추천
    if parsed['filters']:
        filters = build_filters_from_parsed(parsed['filters'])
        count = CylinderRepository.get_cylinder_count(filters=filters)
        
        # 기본 검색 결과
        description_parts = []
        for kw in parsed['parsed_keywords']:
            if kw['type'] == 'status':
                description_parts.append(f"{kw['value']} 상태")
            elif kw['type'] == 'location':
                description_parts.append(f"{kw['value']} 위치")
            elif kw['type'] == 'gas_name':
                description_parts.append(f"{kw['value']} 가스")
            elif kw['type'] == 'pressure':
                description_parts.append("내압관련")
        
        if description_parts:
            suggestions.append({
                'text': query,
                'description': ' + '.join(description_parts),
                'count': count,
                'filters': parsed['filters'],
            })
    
    # 추가 추천
    if 'gas_keyword' in parsed['filters']:
        gas = parsed['filters']['gas_keyword']
        
        # 가스 + 각 상태
        for status in ['보관', '충전', '출하', '이상']:
            test_filters = {'gas_name': gas, 'status': status}
            count = CylinderRepository.get_cylinder_count(filters=test_filters)
            if count > 0:
                suggestions.append({
                    'text': f"{gas} {status}",
                    'description': f"{gas} 가스 중 {status} 상태",
                    'count': count,
                    'filters': {'gas_keyword': gas, 'statuses': [status]},
                })
    
    return JsonResponse({'suggestions': suggestions[:5]})


def apply_scenario(request, scenario_key):
    """시나리오 프리셋 적용 - 리다이렉트"""
    from .search_parser import SCENARIO_PRESETS
    from urllib.parse import urlencode
    
    if scenario_key not in SCENARIO_PRESETS:
        return redirect('cylinders:list')
    
    preset = SCENARIO_PRESETS[scenario_key]
    filters = preset['filters']
    
    # URL 파라미터 구성
    params = {}
    
    if 'statuses' in filters:
        # 다중 상태는 쉼표로 구분
        params['statuses'] = ','.join(filters['statuses'])
    
    if 'pressure_expiring_soon' in filters:
        params['pressure_expiring'] = '1'
        params['pressure_days'] = filters.get('pressure_days', 30)
    
    if 'pressure_expired' in filters:
        params['pressure_expired'] = '1'
    
    if 'days' in filters:
        params['days'] = filters['days']
    
    if 'days_since_event' in filters:
        params['days_since_event'] = filters['days_since_event']
    
    params['scenario'] = scenario_key
    
    url = f"{request.build_absolute_uri('/cylinders/')}?{urlencode(params)}"
    return redirect(url)
