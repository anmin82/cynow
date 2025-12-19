from django.shortcuts import render
from django.http import HttpResponse
from django.utils import timezone
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from core.repositories.cylinder_repository import CylinderRepository
from history.models import HistInventorySnapshot


def weekly_report(request):
    """주간 보고서"""
    # 최근 7일 데이터
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=7)
    
    # 현재 인벤토리
    current_inventory = CylinderRepository.get_inventory_summary()
    
    # 일주일 전 스냅샷
    week_ago_snapshots = HistInventorySnapshot.objects.filter(
        snapshot_datetime__date=start_date,
        snapshot_type='DAILY'
    )
    
    # 현재와 비교를 위한 집계
    current_summary = {}
    week_ago_summary = {}
    
    for row in current_inventory:
        key = f"{row.get('gas_name')}_{row.get('status')}"
        current_summary[key] = row.get('qty', 0)
    
    for snapshot in week_ago_snapshots:
        key = f"{snapshot.gas_name}_{snapshot.status}"
        week_ago_summary[key] = week_ago_summary.get(key, 0) + snapshot.qty
    
    # 비교 데이터 생성
    comparison_data = []
    all_keys = set(current_summary.keys()) | set(week_ago_summary.keys())
    for key in all_keys:
        gas_name, status = key.rsplit('_', 1)
        current_qty = current_summary.get(key, 0)
        week_ago_qty = week_ago_summary.get(key, 0)
        delta = current_qty - week_ago_qty
        comparison_data.append({
            'gas_name': gas_name,
            'status': status,
            'current_qty': current_qty,
            'week_ago_qty': week_ago_qty,
            'delta': delta,
        })
    
    comparison_data.sort(key=lambda x: abs(x['delta']), reverse=True)
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'current_inventory': current_inventory[:50],  # 최대 50개
        'comparison_data': comparison_data[:20],  # Top 20 변동
    }
    return render(request, 'reports/weekly.html', context)


def monthly_report(request):
    """월간 보고서"""
    # 이번 달 데이터
    today = timezone.now().date()
    start_date = today.replace(day=1)  # 이번 달 1일
    end_date = today
    
    # 현재 인벤토리
    current_inventory = CylinderRepository.get_inventory_summary()
    
    # 지난 달 말일 스냅샷
    last_month_end = (start_date - timedelta(days=1))
    last_month_snapshots = HistInventorySnapshot.objects.filter(
        snapshot_datetime__date=last_month_end,
        snapshot_type='DAILY'
    )
    
    # 현재와 비교를 위한 집계
    current_summary = {}
    last_month_summary = {}
    
    for row in current_inventory:
        key = f"{row.get('gas_name')}_{row.get('status')}"
        current_summary[key] = row.get('qty', 0)
    
    for snapshot in last_month_snapshots:
        key = f"{snapshot.gas_name}_{snapshot.status}"
        last_month_summary[key] = last_month_summary.get(key, 0) + snapshot.qty
    
    # 비교 데이터 생성
    comparison_data = []
    all_keys = set(current_summary.keys()) | set(last_month_summary.keys())
    for key in all_keys:
        gas_name, status = key.rsplit('_', 1)
        current_qty = current_summary.get(key, 0)
        last_month_qty = last_month_summary.get(key, 0)
        delta = current_qty - last_month_qty
        comparison_data.append({
            'gas_name': gas_name,
            'status': status,
            'current_qty': current_qty,
            'last_month_qty': last_month_qty,
            'delta': delta,
        })
    
    comparison_data.sort(key=lambda x: abs(x['delta']), reverse=True)
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'current_inventory': current_inventory[:50],  # 최대 50개
        'comparison_data': comparison_data[:20],  # Top 20 변동
    }
    return render(request, 'reports/monthly.html', context)


def export_weekly_excel(request):
    """주간 보고서 Excel 다운로드"""
    # 최근 7일 데이터
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=7)
    
    # 현재 인벤토리
    current_inventory = CylinderRepository.get_inventory_summary()
    
    # 일주일 전 스냅샷
    week_ago_snapshots = HistInventorySnapshot.objects.filter(
        snapshot_datetime__date=start_date,
        snapshot_type='DAILY'
    )
    
    # 현재와 비교를 위한 집계
    current_summary = {}
    week_ago_summary = {}
    
    for row in current_inventory:
        key = f"{row.get('gas_name')}_{row.get('status')}"
        current_summary[key] = row.get('qty', 0)
    
    for snapshot in week_ago_snapshots:
        key = f"{snapshot.gas_name}_{snapshot.status}"
        week_ago_summary[key] = week_ago_summary.get(key, 0) + snapshot.qty
    
    # 비교 데이터 생성
    comparison_data = []
    all_keys = set(current_summary.keys()) | set(week_ago_summary.keys())
    for key in all_keys:
        gas_name, status = key.rsplit('_', 1)
        current_qty = current_summary.get(key, 0)
        week_ago_qty = week_ago_summary.get(key, 0)
        delta = current_qty - week_ago_qty
        comparison_data.append({
            'gas_name': gas_name,
            'status': status,
            'current_qty': current_qty,
            'week_ago_qty': week_ago_qty,
            'delta': delta,
        })
    
    comparison_data.sort(key=lambda x: abs(x['delta']), reverse=True)
    
    # 엑셀 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "주간 보고서"
    
    # 헤더
    headers = ['가스명', '상태', '1주 전', '현재', '변동']
    ws.append(headers)
    
    # 헤더 스타일
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # 데이터
    for item in comparison_data:
        ws.append([
            item['gas_name'],
            item['status'],
            item['week_ago_qty'],
            item['current_qty'],
            item['delta'],
        ])
    
    # 응답 생성
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="cynow_weekly_report_{start_date}_{end_date}.xlsx"'
    wb.save(response)
    
    return response


def export_monthly_excel(request):
    """월간 보고서 Excel 다운로드"""
    # 이번 달 데이터
    today = timezone.now().date()
    start_date = today.replace(day=1)  # 이번 달 1일
    end_date = today
    
    # 현재 인벤토리
    current_inventory = CylinderRepository.get_inventory_summary()
    
    # 지난 달 말일 스냅샷
    last_month_end = (start_date - timedelta(days=1))
    last_month_snapshots = HistInventorySnapshot.objects.filter(
        snapshot_datetime__date=last_month_end,
        snapshot_type='DAILY'
    )
    
    # 현재와 비교를 위한 집계
    current_summary = {}
    last_month_summary = {}
    
    for row in current_inventory:
        key = f"{row.get('gas_name')}_{row.get('status')}"
        current_summary[key] = row.get('qty', 0)
    
    for snapshot in last_month_snapshots:
        key = f"{snapshot.gas_name}_{snapshot.status}"
        last_month_summary[key] = last_month_summary.get(key, 0) + snapshot.qty
    
    # 비교 데이터 생성
    comparison_data = []
    all_keys = set(current_summary.keys()) | set(last_month_summary.keys())
    for key in all_keys:
        gas_name, status = key.rsplit('_', 1)
        current_qty = current_summary.get(key, 0)
        last_month_qty = last_month_summary.get(key, 0)
        delta = current_qty - last_month_qty
        comparison_data.append({
            'gas_name': gas_name,
            'status': status,
            'current_qty': current_qty,
            'last_month_qty': last_month_qty,
            'delta': delta,
        })
    
    comparison_data.sort(key=lambda x: abs(x['delta']), reverse=True)
    
    # 엑셀 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "월간 보고서"
    
    # 헤더
    headers = ['가스명', '상태', '지난 달 말', '현재', '변동']
    ws.append(headers)
    
    # 헤더 스타일
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # 데이터
    for item in comparison_data:
        ws.append([
            item['gas_name'],
            item['status'],
            item['last_month_qty'],
            item['current_qty'],
            item['delta'],
        ])
    
    # 응답 생성
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="cynow_monthly_report_{start_date.strftime("%Y%m")}.xlsx"'
    wb.save(response)
    
    return response
