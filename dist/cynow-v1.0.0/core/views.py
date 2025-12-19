"""CYNOW 전용 관리 페이지"""
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST, require_http_methods
from django.db import connection
from django.core.paginator import Paginator
import csv
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill


def is_staff(user):
    """스태프 권한 확인"""
    return user.is_staff


@login_required
@user_passes_test(is_staff)
def policy_management(request):
    """정책 관리 메인 페이지"""
    return render(request, 'core/policy_management.html')


@login_required
@user_passes_test(is_staff)
def enduser_default_list(request):
    """EndUser 기본값 목록"""
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT 
                id, gas_name, capacity, valve_spec_code, cylinder_spec_code,
                default_enduser, is_active, created_at, updated_at
            FROM cy_enduser_default
            ORDER BY gas_name, capacity NULLS LAST, updated_at DESC
        """)
        columns = [col[0] for col in cursor.description]
        defaults = [dict(zip(columns, row)) for row in cursor.fetchall()]
        
        # 밸브 스펙과 용기 스펙 이름 조회
        for default in defaults:
            # 밸브 스펙 이름 조회 (마스터 테이블 우선, 없으면 스냅샷 테이블)
            if default.get('valve_spec_code'):
                # 먼저 마스터 테이블에서 조회
                cursor.execute("""
                    SELECT "NAME"
                    FROM "fcms_cdc"."ma_valve_specs"
                    WHERE "VALVE_SPEC_CODE" = %s
                    LIMIT 1
                """, [default['valve_spec_code']])
                result = cursor.fetchone()
                if result and result[0]:
                    default['valve_spec_name'] = result[0]
                else:
                    # 마스터 테이블에 없으면 스냅샷 테이블에서 조회
                    cursor.execute("""
                        SELECT DISTINCT raw_valve_spec_name
                        FROM cy_cylinder_current
                        WHERE raw_valve_spec_code = %s
                          AND raw_valve_spec_name IS NOT NULL
                        LIMIT 1
                    """, [default['valve_spec_code']])
                    result = cursor.fetchone()
                    default['valve_spec_name'] = result[0] if result else default['valve_spec_code']
            else:
                default['valve_spec_name'] = None
            
            # 용기 스펙 이름 조회 (마스터 테이블 우선, 없으면 스냅샷 테이블)
            if default.get('cylinder_spec_code'):
                # 먼저 마스터 테이블에서 조회
                cursor.execute("""
                    SELECT "NAME"
                    FROM "fcms_cdc"."ma_cylinder_specs"
                    WHERE "CYLINDER_SPEC_CODE" = %s
                    LIMIT 1
                """, [default['cylinder_spec_code']])
                result = cursor.fetchone()
                if result and result[0]:
                    default['cylinder_spec_name'] = result[0]
                else:
                    # 마스터 테이블에 없으면 스냅샷 테이블에서 조회
                    cursor.execute("""
                        SELECT DISTINCT raw_cylinder_spec_name
                        FROM cy_cylinder_current
                        WHERE raw_cylinder_spec_code = %s
                          AND raw_cylinder_spec_name IS NOT NULL
                        LIMIT 1
                    """, [default['cylinder_spec_code']])
                    result = cursor.fetchone()
                    default['cylinder_spec_name'] = result[0] if result else default['cylinder_spec_code']
            else:
                default['cylinder_spec_name'] = None
    
    # 페이지네이션
    paginator = Paginator(defaults, 50)
    page = request.GET.get('page', 1)
    try:
        page_obj = paginator.page(page)
    except:
        page_obj = paginator.page(1)
    
    context = {
        'defaults': page_obj,
    }
    return render(request, 'core/enduser_default_list.html', context)


@login_required
@user_passes_test(is_staff)
@require_http_methods(["GET", "POST"])
def enduser_default_add(request):
    """EndUser 기본값 추가"""
    if request.method == 'POST':
        gas_name = request.POST.get('gas_name', '').strip()
        capacity = request.POST.get('capacity', '').strip() or None
        valve_spec_code = request.POST.get('valve_spec_code', '').strip() or None
        cylinder_spec_code = request.POST.get('cylinder_spec_code', '').strip() or None
        default_enduser = request.POST.get('default_enduser', 'SDC').strip()
        
        if not gas_name or not default_enduser:
            messages.error(request, '가스명과 기본 EndUser는 필수입니다.')
            return redirect('core:enduser_default_add')
        
        with connection.cursor() as cursor:
            try:
                cursor.execute("""
                    INSERT INTO cy_enduser_default 
                    (gas_name, capacity, valve_spec_code, cylinder_spec_code, default_enduser, is_active)
                    VALUES (%s, %s, %s, %s, %s, TRUE)
                    ON CONFLICT (gas_name, capacity, valve_spec_code, cylinder_spec_code) 
                    DO UPDATE SET 
                        default_enduser = EXCLUDED.default_enduser,
                        is_active = TRUE,
                        updated_at = NOW()
                """, [gas_name, capacity, valve_spec_code, cylinder_spec_code, default_enduser])
                
                messages.success(request, f'기본값이 추가되었습니다: {gas_name} → {default_enduser}')
                
                # 스냅샷 갱신 (해당 조건의 모든 용기)
                cursor.execute("""
                    SELECT DISTINCT cylinder_no
                    FROM cy_cylinder_current
                    WHERE raw_gas_name = %s
                      AND (%s IS NULL OR raw_capacity = %s)
                      AND (%s IS NULL OR raw_valve_spec_code = %s)
                      AND (%s IS NULL OR raw_cylinder_spec_code = %s)
                """, [gas_name, capacity, capacity, valve_spec_code, valve_spec_code, 
                      cylinder_spec_code, cylinder_spec_code])
                for row in cursor.fetchall():
                    cursor.execute("SELECT sync_cylinder_current_single(%s)", [row[0]])
                
                return redirect('core:enduser_default_list')
            except Exception as e:
                messages.error(request, f'오류: {str(e)}')
    
    # GET 요청 시 가스명 목록 조회 (자동완성용)
    # cy_cylinder_current 테이블에서 가스명 가져오기
    with connection.cursor() as cursor:
        try:
            cursor.execute("""
                SELECT DISTINCT dashboard_gas_name
                FROM cy_cylinder_current
                WHERE dashboard_gas_name IS NOT NULL
                  AND dashboard_gas_name != ''
                ORDER BY dashboard_gas_name
                LIMIT 100
            """)
            gas_names = [row[0] for row in cursor.fetchall()]
        except Exception:
            # 테이블이 없거나 오류 발생 시 빈 리스트 반환
            gas_names = []
        
        # EndUser 목록 조회
        try:
            cursor.execute("""
                SELECT enduser_code, enduser_name
                FROM cy_enduser_master
                WHERE is_active = TRUE
                ORDER BY enduser_code
            """)
            endusers = [{'code': row[0], 'name': row[1]} for row in cursor.fetchall()]
        except Exception:
            endusers = []
    
    import json
    context = {
        'gas_names': json.dumps(gas_names, ensure_ascii=False),
        'endusers': endusers,
    }
    return render(request, 'core/enduser_default_add.html', context)


@login_required
@user_passes_test(is_staff)
@require_http_methods(["GET", "POST"])
def enduser_default_edit(request, default_id):
    """EndUser 기본값 수정"""
    with connection.cursor() as cursor:
        # 기존 데이터 조회
        cursor.execute("""
            SELECT 
                id, gas_name, capacity, valve_spec_code, cylinder_spec_code,
                default_enduser, is_active
            FROM cy_enduser_default
            WHERE id = %s
        """, [default_id])
        row = cursor.fetchone()
        
        if not row:
            messages.error(request, '기본값을 찾을 수 없습니다.')
            return redirect('core:enduser_default_list')
        
        columns = [col[0] for col in cursor.description]
        default_data = dict(zip(columns, row))
        
        # capacity를 문자열로 변환 (Decimal 타입 처리)
        if default_data.get('capacity') is not None:
            default_data['capacity'] = str(default_data['capacity'])
        
        # 기존 밸브/용기 스펙 이름 조회
        if default_data.get('valve_spec_code'):
            cursor.execute("""
                SELECT DISTINCT raw_valve_spec_name
                FROM cy_cylinder_current
                WHERE raw_valve_spec_code = %s
                LIMIT 1
            """, [default_data['valve_spec_code']])
            result = cursor.fetchone()
            default_data['valve_spec_name'] = result[0] if result else ''
        else:
            default_data['valve_spec_name'] = ''
        
        if default_data.get('cylinder_spec_code'):
            cursor.execute("""
                SELECT DISTINCT raw_cylinder_spec_name
                FROM cy_cylinder_current
                WHERE raw_cylinder_spec_code = %s
                LIMIT 1
            """, [default_data['cylinder_spec_code']])
            result = cursor.fetchone()
            default_data['cylinder_spec_name'] = result[0] if result else ''
        else:
            default_data['cylinder_spec_name'] = ''
    
    if request.method == 'POST':
        gas_name = request.POST.get('gas_name', '').strip()
        capacity = request.POST.get('capacity', '').strip() or None
        valve_spec_code = request.POST.get('valve_spec_code', '').strip() or None
        cylinder_spec_code = request.POST.get('cylinder_spec_code', '').strip() or None
        default_enduser = request.POST.get('default_enduser', 'SDC').strip()
        is_active = request.POST.get('is_active') == 'on'
        
        if not gas_name or not default_enduser:
            messages.error(request, '가스명과 기본 EndUser는 필수입니다.')
            return redirect('core:enduser_default_edit', default_id=default_id)
        
        with connection.cursor() as cursor:
            try:
                # 기존 조건과 새 조건이 다르면 unique constraint 위반 가능성 체크
                old_key = (default_data['gas_name'], default_data['capacity'], 
                          default_data['valve_spec_code'], default_data['cylinder_spec_code'])
                new_key = (gas_name, capacity, valve_spec_code, cylinder_spec_code)
                
                if old_key != new_key:
                    # 다른 레코드와 충돌하는지 확인
                    cursor.execute("""
                        SELECT id FROM cy_enduser_default
                        WHERE gas_name = %s
                          AND (%s IS NULL AND capacity IS NULL OR capacity = %s)
                          AND (%s IS NULL AND valve_spec_code IS NULL OR valve_spec_code = %s)
                          AND (%s IS NULL AND cylinder_spec_code IS NULL OR cylinder_spec_code = %s)
                          AND id != %s
                    """, [gas_name, capacity, capacity, valve_spec_code, valve_spec_code,
                          cylinder_spec_code, cylinder_spec_code, default_id])
                    if cursor.fetchone():
                        messages.error(request, '같은 조건의 기본값이 이미 존재합니다.')
                        return redirect('core:enduser_default_edit', default_id=default_id)
                
                cursor.execute("""
                    UPDATE cy_enduser_default
                    SET gas_name = %s,
                        capacity = %s,
                        valve_spec_code = %s,
                        cylinder_spec_code = %s,
                        default_enduser = %s,
                        is_active = %s,
                        updated_at = NOW()
                    WHERE id = %s
                """, [gas_name, capacity, valve_spec_code, cylinder_spec_code, 
                      default_enduser, is_active, default_id])
                
                messages.success(request, f'기본값이 수정되었습니다: {gas_name} → {default_enduser}')
                
                # 스냅샷 갱신 (기존 조건과 새 조건 모두)
                for gas, cap, valve, cyl in [old_key, new_key]:
                    cursor.execute("""
                        SELECT DISTINCT cylinder_no
                        FROM cy_cylinder_current
                        WHERE raw_gas_name = %s
                          AND (%s IS NULL OR raw_capacity = %s)
                          AND (%s IS NULL OR raw_valve_spec_code = %s)
                          AND (%s IS NULL OR raw_cylinder_spec_code = %s)
                    """, [gas, cap, cap, valve, valve, cyl, cyl])
                    for row in cursor.fetchall():
                        cursor.execute("SELECT sync_cylinder_current_single(%s)", [row[0]])
                
                return redirect('core:enduser_default_list')
            except Exception as e:
                messages.error(request, f'오류: {str(e)}')
    
    # GET 요청 시 가스명 목록 조회 (자동완성용)
    # cy_cylinder_current 테이블에서 가스명 가져오기
    with connection.cursor() as cursor:
        try:
            cursor.execute("""
                SELECT DISTINCT dashboard_gas_name
                FROM cy_cylinder_current
                WHERE dashboard_gas_name IS NOT NULL
                  AND dashboard_gas_name != ''
                ORDER BY dashboard_gas_name
                LIMIT 100
            """)
            gas_names = [row[0] for row in cursor.fetchall()]
        except Exception:
            # 테이블이 없거나 오류 발생 시 빈 리스트 반환
            gas_names = []
        
        # EndUser 목록 조회
        try:
            cursor.execute("""
                SELECT enduser_code, enduser_name
                FROM cy_enduser_master
                WHERE is_active = TRUE
                ORDER BY enduser_code
            """)
            endusers = [{'code': row[0], 'name': row[1]} for row in cursor.fetchall()]
        except Exception:
            endusers = []
    
    import json
    context = {
        'default': default_data,
        'gas_names': json.dumps(gas_names, ensure_ascii=False),
        'endusers': endusers,
    }
    return render(request, 'core/enduser_default_edit.html', context)


@login_required
@user_passes_test(is_staff)
@require_POST
def enduser_default_toggle(request, default_id):
    """EndUser 기본값 활성화/비활성화"""
    with connection.cursor() as cursor:
        cursor.execute("""
            UPDATE cy_enduser_default
            SET is_active = NOT is_active, updated_at = NOW()
            WHERE id = %s
            RETURNING gas_name, capacity, valve_spec_code, cylinder_spec_code, is_active
        """, [default_id])
        result = cursor.fetchone()
        
        if result:
            gas_name, capacity, valve_spec_code, cylinder_spec_code, is_active = result
            status = '활성화' if is_active else '비활성화'
            messages.success(request, f'{gas_name} 기본값이 {status}되었습니다.')
            
            # 스냅샷 갱신
            cursor.execute("""
                SELECT DISTINCT cylinder_no
                FROM cy_cylinder_current
                WHERE raw_gas_name = %s
                  AND (%s IS NULL OR raw_capacity = %s)
                  AND (%s IS NULL OR raw_valve_spec_code = %s)
                  AND (%s IS NULL OR raw_cylinder_spec_code = %s)
            """, [gas_name, capacity, capacity, valve_spec_code, valve_spec_code, 
                  cylinder_spec_code, cylinder_spec_code])
            for row in cursor.fetchall():
                cursor.execute("SELECT sync_cylinder_current_single(%s)", [row[0]])
    
    return redirect('core:enduser_default_list')


@login_required
@user_passes_test(is_staff)
@require_POST
def enduser_default_delete(request, default_id):
    """EndUser 기본값 삭제"""
    with connection.cursor() as cursor:
        # 삭제 전 정보 조회
        cursor.execute("""
            SELECT gas_name, capacity, valve_spec_code, cylinder_spec_code
            FROM cy_enduser_default
            WHERE id = %s
        """, [default_id])
        result = cursor.fetchone()
        
        if not result:
            messages.error(request, '기본값을 찾을 수 없습니다.')
            return redirect('core:enduser_default_list')
        
        gas_name, capacity, valve_spec_code, cylinder_spec_code = result
        
        # 삭제 실행
        cursor.execute("""
            DELETE FROM cy_enduser_default
            WHERE id = %s
        """, [default_id])
        
        messages.success(request, f'기본값이 삭제되었습니다: {gas_name}')
        
        # 스냅샷 갱신 (삭제된 정책이 적용되던 용기들)
        cursor.execute("""
            SELECT DISTINCT cylinder_no
            FROM cy_cylinder_current
            WHERE raw_gas_name = %s
              AND (%s IS NULL OR raw_capacity = %s)
              AND (%s IS NULL OR raw_valve_spec_code = %s)
              AND (%s IS NULL OR raw_cylinder_spec_code = %s)
        """, [gas_name, capacity, capacity, valve_spec_code, valve_spec_code, 
              cylinder_spec_code, cylinder_spec_code])
        for row in cursor.fetchall():
            cursor.execute("SELECT sync_cylinder_current_single(%s)", [row[0]])
    
    return redirect('core:enduser_default_list')


@login_required
@user_passes_test(is_staff)
def enduser_exception_list(request):
    """EndUser 예외 목록"""
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT 
                id, cylinder_no, enduser, reason, is_active, 
                created_at, updated_at
            FROM cy_enduser_exception
            ORDER BY updated_at DESC
        """)
        columns = [col[0] for col in cursor.description]
        exceptions = [dict(zip(columns, row)) for row in cursor.fetchall()]
    
    # 페이지네이션
    paginator = Paginator(exceptions, 50)
    page = request.GET.get('page', 1)
    try:
        page_obj = paginator.page(page)
    except:
        page_obj = paginator.page(1)
    
    context = {
        'exceptions': page_obj,
    }
    return render(request, 'core/enduser_exception_list.html', context)


@login_required
@user_passes_test(is_staff)
@require_http_methods(["GET", "POST"])
def enduser_exception_add(request):
    """EndUser 예외 추가"""
    if request.method == 'POST':
        cylinder_no = request.POST.get('cylinder_no', '').strip().rstrip()  # 오른쪽 공백 제거
        enduser = request.POST.get('enduser', '').strip()
        reason = request.POST.get('reason', '').strip()
        
        if not cylinder_no or not enduser:
            messages.error(request, '용기번호와 EndUser는 필수입니다.')
            return redirect('core:enduser_exception_add')
        
        with connection.cursor() as cursor:
            try:
                # 용기번호 오른쪽 공백 제거 후 저장
                cylinder_no_trimmed = cylinder_no.rstrip()
                cursor.execute("""
                    INSERT INTO cy_enduser_exception 
                    (cylinder_no, enduser, reason, is_active)
                    VALUES (%s, %s, %s, TRUE)
                    ON CONFLICT (cylinder_no) 
                    DO UPDATE SET 
                        enduser = EXCLUDED.enduser,
                        reason = EXCLUDED.reason,
                        is_active = TRUE,
                        updated_at = NOW()
                """, [cylinder_no_trimmed, enduser, reason])
                
                messages.success(request, f'예외가 추가되었습니다: {cylinder_no_trimmed} → {enduser}')
                
                # 스냅샷 갱신 (TRIM된 용기번호로)
                cursor.execute("SELECT sync_cylinder_current_single(%s)", [cylinder_no_trimmed])
                
                return redirect('core:enduser_exception_list')
            except Exception as e:
                messages.error(request, f'오류: {str(e)}')
    
    # EndUser 목록 조회
    with connection.cursor() as cursor:
        try:
            cursor.execute("""
                SELECT enduser_code, enduser_name
                FROM cy_enduser_master
                WHERE is_active = TRUE
                ORDER BY enduser_code
            """)
            endusers = [{'code': row[0], 'name': row[1]} for row in cursor.fetchall()]
        except Exception:
            endusers = []
    
    context = {
        'endusers': endusers,
    }
    return render(request, 'core/enduser_exception_add.html', context)


@login_required
@user_passes_test(is_staff)
@require_POST
def enduser_exception_upload(request):
    """EndUser 예외 Excel/CSV 업로드"""
    if 'excel_file' not in request.FILES:
        messages.error(request, 'Excel 파일을 선택해주세요.')
        return redirect('core:enduser_exception_list')
    
    excel_file = request.FILES['excel_file']
    file_extension = excel_file.name.split('.')[-1].lower()
    
    count = 0
    errors = []
    
    try:
        with connection.cursor() as cursor:
            if file_extension == 'xlsx' or file_extension == 'xls':
                # Excel 파일 읽기
                try:
                    wb = load_workbook(excel_file, data_only=True)
                    ws = wb.active
                    
                    # 헤더 확인 (첫 번째 행)
                    headers = [cell.value for cell in ws[1]]
                    header_map = {}
                    for idx, header in enumerate(headers, 1):
                        if header:
                            header_lower = str(header).lower().strip()
                            if 'cylinder' in header_lower or '용기' in header_lower:
                                header_map['cylinder_no'] = idx
                            elif 'enduser' in header_lower or '엔드유저' in header_lower:
                                header_map['enduser'] = idx
                            elif 'reason' in header_lower or '사유' in header_lower:
                                header_map['reason'] = idx
                    
                    if 'cylinder_no' not in header_map or 'enduser' not in header_map:
                        messages.error(request, 'Excel 파일 형식이 올바르지 않습니다. 템플릿 파일을 다운로드하여 사용하세요.')
                        return redirect('core:enduser_exception_list')
                    
                    # 데이터 행 처리 (2번째 행부터)
                    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=False), 2):
                        cylinder_no = str(row[header_map['cylinder_no'] - 1].value or '').strip().rstrip()  # 오른쪽 공백 제거
                        enduser = str(row[header_map['enduser'] - 1].value or '').strip()
                        reason = str(row[header_map.get('reason', 0) - 1].value or '').strip() if header_map.get('reason') else ''
                        
                        if not cylinder_no or not enduser:
                            if cylinder_no or enduser:  # 일부만 있는 경우에만 오류로 표시
                                errors.append(f'{row_num}행: 용기번호 또는 EndUser가 없습니다.')
                            continue
                        
                        try:
                            # 용기번호 오른쪽 공백 제거 후 저장
                            cylinder_no_trimmed = cylinder_no.rstrip()
                            cursor.execute("""
                                INSERT INTO cy_enduser_exception 
                                (cylinder_no, enduser, reason, is_active)
                                VALUES (%s, %s, %s, TRUE)
                                ON CONFLICT (cylinder_no) 
                                DO UPDATE SET 
                                    enduser = EXCLUDED.enduser,
                                    reason = EXCLUDED.reason,
                                    is_active = TRUE,
                                    updated_at = NOW()
                            """, [cylinder_no_trimmed, enduser, reason])
                            count += 1
                            
                            # 스냅샷 갱신 (TRIM된 용기번호로)
                            cursor.execute("SELECT sync_cylinder_current_single(%s)", [cylinder_no_trimmed])
                        except Exception as e:
                            errors.append(f'{row_num}행 ({cylinder_no}): {str(e)}')
                            
                except Exception as e:
                    messages.error(request, f'Excel 파일 처리 오류: {str(e)}')
                    return redirect('core:enduser_exception_list')
                    
            elif file_extension == 'csv':
                # CSV 파일 읽기 (기존 로직 유지)
                decoded_file = excel_file.read().decode('utf-8-sig')
                io_string = io.StringIO(decoded_file)
                reader = csv.DictReader(io_string)
                
                for i, row in enumerate(reader, 1):
                    cylinder_no = row.get('cylinder_no', '').strip().rstrip()  # 오른쪽 공백 제거
                    enduser = row.get('enduser', '').strip()
                    reason = row.get('reason', '').strip()
                    
                    if not cylinder_no or not enduser:
                        errors.append(f'{i}행: 용기번호 또는 EndUser가 없습니다.')
                        continue
                    
                    try:
                        # 용기번호 오른쪽 공백 제거 후 저장
                        cylinder_no_trimmed = cylinder_no.rstrip()
                        cursor.execute("""
                            INSERT INTO cy_enduser_exception 
                            (cylinder_no, enduser, reason, is_active)
                            VALUES (%s, %s, %s, TRUE)
                            ON CONFLICT (cylinder_no) 
                            DO UPDATE SET 
                                enduser = EXCLUDED.enduser,
                                reason = EXCLUDED.reason,
                                is_active = TRUE,
                                updated_at = NOW()
                        """, [cylinder_no_trimmed, enduser, reason])
                        count += 1
                        
                        # 스냅샷 갱신 (TRIM된 용기번호로)
                        cursor.execute("SELECT sync_cylinder_current_single(%s)", [cylinder_no_trimmed])
                    except Exception as e:
                        errors.append(f'{i}행 ({cylinder_no}): {str(e)}')
            else:
                messages.error(request, '지원하지 않는 파일 형식입니다. Excel(.xlsx) 또는 CSV(.csv) 파일을 사용하세요.')
                return redirect('core:enduser_exception_list')
        
        if count > 0:
            messages.success(request, f'{count}개 예외가 추가되었습니다.')
        if errors:
            messages.warning(request, f'{len(errors)}개 오류가 발생했습니다.')
            for error in errors[:10]:  # 최대 10개만 표시
                messages.warning(request, error)
        
    except Exception as e:
        messages.error(request, f'파일 처리 오류: {str(e)}')
    
    return redirect('core:enduser_exception_list')


@login_required
@user_passes_test(is_staff)
def enduser_exception_template(request):
    """EndUser 예외 Excel 템플릿 다운로드"""
    wb = Workbook()
    ws = wb.active
    ws.title = "EndUser 예외"
    
    # 헤더 스타일
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 헤더 작성
    headers = ['용기번호', 'EndUser', '사유']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 예시 데이터 (2행)
    example_data = [
        ['CYL001', 'LGD', 'LGD 전용 용기'],
        ['CYL002', 'SDC', ''],
    ]
    for row_num, row_data in enumerate(example_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # 열 너비 조정
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    
    # 응답 생성
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="enduser_exception_template.xlsx"'
    
    wb.save(response)
    return response


@login_required
@user_passes_test(is_staff)
@require_POST
def enduser_exception_toggle(request, exception_id):
    """EndUser 예외 활성화/비활성화"""
    with connection.cursor() as cursor:
        cursor.execute("""
            UPDATE cy_enduser_exception
            SET is_active = NOT is_active, updated_at = NOW()
            WHERE id = %s
            RETURNING cylinder_no, is_active
        """, [exception_id])
        result = cursor.fetchone()
        
        if result:
            cylinder_no, is_active = result
            status = '활성화' if is_active else '비활성화'
            messages.success(request, f'{cylinder_no}가 {status}되었습니다.')
            
            # 스냅샷 갱신
            cursor.execute("SELECT sync_cylinder_current_single(%s)", [cylinder_no])
    
    return redirect('core:enduser_exception_list')


@login_required
@user_passes_test(is_staff)
def valve_group_list(request):
    """밸브 그룹 목록"""
    with connection.cursor() as cursor:
        # 그룹 목록
        cursor.execute("""
            SELECT 
                id, group_name, description, is_active,
                created_at, updated_at
            FROM cy_valve_group
            ORDER BY updated_at DESC
        """)
        columns = [col[0] for col in cursor.description]
        groups = [dict(zip(columns, row)) for row in cursor.fetchall()]
        
        # 각 그룹의 매핑 수
        for group in groups:
            cursor.execute("""
                SELECT COUNT(*) 
                FROM cy_valve_group_mapping 
                WHERE group_id = %s AND is_active = TRUE
            """, [group['id']])
            group['mapping_count'] = cursor.fetchone()[0]
    
    context = {
        'groups': groups,
    }
    return render(request, 'core/valve_group_list.html', context)


@login_required
@user_passes_test(is_staff)
@require_http_methods(["GET", "POST"])
def valve_group_add(request):
    """밸브 그룹 추가"""
    if request.method == 'POST':
        group_name = request.POST.get('group_name', '').strip()
        description = request.POST.get('description', '').strip()
        
        if not group_name:
            messages.error(request, '그룹명은 필수입니다.')
            return redirect('core:valve_group_add')
        
        with connection.cursor() as cursor:
            try:
                cursor.execute("""
                    INSERT INTO cy_valve_group 
                    (group_name, description, is_active)
                    VALUES (%s, %s, TRUE)
                    ON CONFLICT (group_name) 
                    DO UPDATE SET 
                        description = EXCLUDED.description,
                        is_active = TRUE,
                        updated_at = NOW()
                    RETURNING id
                """, [group_name, description])
                
                group_id = cursor.fetchone()[0]
                messages.success(request, f'밸브 그룹이 생성되었습니다: {group_name}')
                return redirect('core:valve_group_detail', group_id=group_id)
            except Exception as e:
                messages.error(request, f'오류: {str(e)}')
    
    return render(request, 'core/valve_group_add.html')


@login_required
@user_passes_test(is_staff)
def valve_group_detail(request, group_id):
    """밸브 그룹 상세 (매핑 관리)"""
    with connection.cursor() as cursor:
        # 그룹 정보
        cursor.execute("""
            SELECT id, group_name, description, is_active
            FROM cy_valve_group
            WHERE id = %s
        """, [group_id])
        group = dict(zip([col[0] for col in cursor.description], cursor.fetchone()))
        
        # 매핑 목록
        cursor.execute("""
            SELECT 
                id, valve_spec_code, valve_spec_name, 
                is_primary, is_active, created_at, updated_at
            FROM cy_valve_group_mapping
            WHERE group_id = %s
            ORDER BY is_primary DESC, valve_spec_name
        """, [group_id])
        mappings = [dict(zip([col[0] for col in cursor.description], row)) 
                   for row in cursor.fetchall()]
        
        # 사용 가능한 밸브 스펙 목록 (검색용)
        search_query = request.GET.get('search', '')
        available_valves = []
        if search_query:
            cursor.execute("""
                SELECT DISTINCT 
                    vs."VALVE_SPEC_CODE",
                    vs."NAME"
                FROM "fcms_cdc"."ma_valve_specs" vs
                WHERE vs."NAME" ILIKE %s
                  AND NOT EXISTS (
                      SELECT 1 FROM cy_valve_group_mapping vgm
                      WHERE vgm.valve_spec_code = vs."VALVE_SPEC_CODE"
                        AND vgm.valve_spec_name = vs."NAME"
                        AND vgm.is_active = TRUE
                  )
                ORDER BY vs."NAME"
                LIMIT 20
            """, [f'%{search_query}%'])
            available_valves = [dict(zip(['code', 'name'], row)) 
                              for row in cursor.fetchall()]
    
    context = {
        'group': group,
        'mappings': mappings,
        'available_valves': available_valves,
        'search_query': search_query,
    }
    return render(request, 'core/valve_group_detail.html', context)


@login_required
@user_passes_test(is_staff)
@require_POST
def valve_group_mapping_add(request, group_id):
    """밸브 그룹 매핑 추가"""
    valve_spec_code = request.POST.get('valve_spec_code', '').strip()
    valve_spec_name = request.POST.get('valve_spec_name', '').strip()
    is_primary = request.POST.get('is_primary') == 'on'
    
    if not valve_spec_code or not valve_spec_name:
        messages.error(request, '밸브 스펙 코드와 이름은 필수입니다.')
        return redirect('core:valve_group_detail', group_id=group_id)
    
    with connection.cursor() as cursor:
        try:
            # primary가 이미 있으면 해제
            if is_primary:
                cursor.execute("""
                    UPDATE cy_valve_group_mapping
                    SET is_primary = FALSE
                    WHERE group_id = %s
                """, [group_id])
            
            cursor.execute("""
                INSERT INTO cy_valve_group_mapping 
                (valve_spec_code, valve_spec_name, group_id, is_primary, is_active)
                VALUES (%s, %s, %s, %s, TRUE)
                ON CONFLICT (valve_spec_code, valve_spec_name) 
                DO UPDATE SET 
                    group_id = EXCLUDED.group_id,
                    is_primary = EXCLUDED.is_primary,
                    is_active = TRUE,
                    updated_at = NOW()
            """, [valve_spec_code, valve_spec_name, group_id, is_primary])
            
            messages.success(request, '밸브 매핑이 추가되었습니다.')
            
            # 스냅샷 갱신 (해당 밸브를 사용하는 모든 용기)
            cursor.execute("""
                SELECT DISTINCT cylinder_no
                FROM cy_cylinder_current
                WHERE raw_valve_spec_code = %s
            """, [valve_spec_code])
            for row in cursor.fetchall():
                cursor.execute("SELECT sync_cylinder_current_single(%s)", [row[0]])
            
        except Exception as e:
            messages.error(request, f'오류: {str(e)}')
    
    return redirect('core:valve_group_detail', group_id=group_id)


@login_required
@user_passes_test(is_staff)
@require_POST
def valve_group_mapping_toggle_primary(request, mapping_id):
    """밸브 그룹 매핑 대표 밸브 설정"""
    with connection.cursor() as cursor:
        # 그룹 ID 조회
        cursor.execute("SELECT group_id FROM cy_valve_group_mapping WHERE id = %s", [mapping_id])
        result = cursor.fetchone()
        if not result:
            messages.error(request, '매핑을 찾을 수 없습니다.')
            return redirect('core:valve_group_list')
        
        group_id = result[0]
        
        # 기존 primary 해제
        cursor.execute("""
            UPDATE cy_valve_group_mapping
            SET is_primary = FALSE
            WHERE group_id = %s
        """, [group_id])
        
        # 새로운 primary 설정
        cursor.execute("""
            UPDATE cy_valve_group_mapping
            SET is_primary = TRUE, updated_at = NOW()
            WHERE id = %s
        """, [mapping_id])
        
        messages.success(request, '대표 밸브가 변경되었습니다.')
        
        # 스냅샷 갱신
        cursor.execute("""
            SELECT DISTINCT cylinder_no
            FROM cy_cylinder_current
            WHERE raw_valve_spec_code = (
                SELECT valve_spec_code FROM cy_valve_group_mapping WHERE id = %s
            )
        """, [mapping_id])
        for row in cursor.fetchall():
            cursor.execute("SELECT sync_cylinder_current_single(%s)", [row[0]])
    
    return redirect('core:valve_group_detail', group_id=group_id)


@login_required
@user_passes_test(is_staff)
@require_POST
def valve_group_mapping_toggle_active(request, mapping_id):
    """밸브 그룹 매핑 활성화/비활성화"""
    with connection.cursor() as cursor:
        cursor.execute("""
            UPDATE cy_valve_group_mapping
            SET is_active = NOT is_active, updated_at = NOW()
            WHERE id = %s
            RETURNING group_id, valve_spec_code
        """, [mapping_id])
        result = cursor.fetchone()
        
        if result:
            group_id, valve_spec_code = result
            messages.success(request, '매핑 상태가 변경되었습니다.')
            
            # 스냅샷 갱신
            cursor.execute("""
                SELECT DISTINCT cylinder_no
                FROM cy_cylinder_current
                WHERE raw_valve_spec_code = %s
            """, [valve_spec_code])
            for row in cursor.fetchall():
                cursor.execute("SELECT sync_cylinder_current_single(%s)", [row[0]])
            
            return redirect('core:valve_group_detail', group_id=group_id)
    
    return redirect('core:valve_group_list')


@login_required
@user_passes_test(is_staff)
def valve_spec_search(request):
    """밸브 스펙 검색 (AJAX)"""
    query = request.GET.get('q', '').strip()
    
    if not query:
        return JsonResponse({'results': []})
    
    with connection.cursor() as cursor:
        try:
            # 먼저 fcms_cdc.ma_valve_specs 테이블에서 검색 시도
            cursor.execute("""
                SELECT DISTINCT 
                    vs."VALVE_SPEC_CODE" as code,
                    vs."NAME" as name
                FROM "fcms_cdc"."ma_valve_specs" vs
                WHERE vs."NAME" ILIKE %s
                ORDER BY vs."NAME"
                LIMIT 20
            """, [f'%{query}%'])
            results = [{'code': row[0], 'name': row[1]} for row in cursor.fetchall()]
        except Exception:
            # 테이블이 없으면 cy_cylinder_current에서 검색
            cursor.execute("""
                SELECT DISTINCT 
                    raw_valve_spec_code as code,
                    raw_valve_spec_name as name
                FROM cy_cylinder_current
                WHERE raw_valve_spec_name ILIKE %s
                  AND raw_valve_spec_code IS NOT NULL
                  AND raw_valve_spec_name IS NOT NULL
                ORDER BY raw_valve_spec_name
                LIMIT 20
            """, [f'%{query}%'])
            results = [{'code': row[0], 'name': row[1]} for row in cursor.fetchall()]
    
    return JsonResponse({'results': results})


@login_required
@user_passes_test(is_staff)
def cylinder_spec_search(request):
    """용기 스펙 검색 (AJAX)"""
    query = request.GET.get('q', '').strip()
    
    if not query:
        return JsonResponse({'results': []})
    
    with connection.cursor() as cursor:
        try:
            # 먼저 fcms_cdc.ma_cylinder_specs 테이블에서 검색 시도
            cursor.execute("""
                SELECT DISTINCT 
                    cs."CYLINDER_SPEC_CODE" as code,
                    cs."NAME" as name
                FROM "fcms_cdc"."ma_cylinder_specs" cs
                WHERE cs."NAME" ILIKE %s
                ORDER BY cs."NAME"
                LIMIT 20
            """, [f'%{query}%'])
            results = [{'code': row[0], 'name': row[1]} for row in cursor.fetchall()]
        except Exception:
            # 테이블이 없으면 cy_cylinder_current에서 검색
            cursor.execute("""
                SELECT DISTINCT 
                    raw_cylinder_spec_code as code,
                    raw_cylinder_spec_name as name
                FROM cy_cylinder_current
                WHERE raw_cylinder_spec_name ILIKE %s
                  AND raw_cylinder_spec_code IS NOT NULL
                  AND raw_cylinder_spec_name IS NOT NULL
                ORDER BY raw_cylinder_spec_name
                LIMIT 20
            """, [f'%{query}%'])
            results = [{'code': row[0], 'name': row[1]} for row in cursor.fetchall()]
    
    return JsonResponse({'results': results})


# EndUser 마스터 관리
@login_required
@user_passes_test(is_staff)
def enduser_master_list(request):
    """EndUser 마스터 목록"""
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT 
                id, enduser_code, enduser_name, description, is_active,
                created_at, updated_at
            FROM cy_enduser_master
            ORDER BY enduser_code
        """)
        columns = [col[0] for col in cursor.description]
        endusers = [dict(zip(columns, row)) for row in cursor.fetchall()]
    
    context = {
        'endusers': endusers,
    }
    return render(request, 'core/enduser_master_list.html', context)


@login_required
@user_passes_test(is_staff)
@require_http_methods(["GET", "POST"])
def enduser_master_add(request):
    """EndUser 마스터 추가"""
    if request.method == 'POST':
        enduser_code = request.POST.get('enduser_code', '').strip().upper()
        enduser_name = request.POST.get('enduser_name', '').strip()
        description = request.POST.get('description', '').strip() or None
        
        if not enduser_code or not enduser_name:
            messages.error(request, 'EndUser 코드와 이름은 필수입니다.')
            return redirect('core:enduser_master_add')
        
        with connection.cursor() as cursor:
            try:
                cursor.execute("""
                    INSERT INTO cy_enduser_master 
                    (enduser_code, enduser_name, description, is_active)
                    VALUES (%s, %s, %s, TRUE)
                    ON CONFLICT (enduser_code) 
                    DO UPDATE SET 
                        enduser_name = EXCLUDED.enduser_name,
                        description = EXCLUDED.description,
                        is_active = TRUE,
                        updated_at = NOW()
                """, [enduser_code, enduser_name, description])
                
                messages.success(request, f'EndUser가 추가되었습니다: {enduser_code}')
                return redirect('core:enduser_master_list')
            except Exception as e:
                messages.error(request, f'오류: {str(e)}')
    
    return render(request, 'core/enduser_master_add.html')


@login_required
@user_passes_test(is_staff)
@require_http_methods(["GET", "POST"])
def enduser_master_edit(request, enduser_id):
    """EndUser 마스터 수정"""
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT id, enduser_code, enduser_name, description, is_active
            FROM cy_enduser_master
            WHERE id = %s
        """, [enduser_id])
        row = cursor.fetchone()
        
        if not row:
            messages.error(request, 'EndUser를 찾을 수 없습니다.')
            return redirect('core:enduser_master_list')
        
        columns = [col[0] for col in cursor.description]
        enduser_data = dict(zip(columns, row))
    
    if request.method == 'POST':
        enduser_code = request.POST.get('enduser_code', '').strip().upper()
        enduser_name = request.POST.get('enduser_name', '').strip()
        description = request.POST.get('description', '').strip() or None
        is_active = request.POST.get('is_active') == 'on'
        
        if not enduser_code or not enduser_name:
            messages.error(request, 'EndUser 코드와 이름은 필수입니다.')
            return redirect('core:enduser_master_edit', enduser_id=enduser_id)
        
        with connection.cursor() as cursor:
            try:
                # 코드 변경 시 중복 체크
                if enduser_code != enduser_data['enduser_code']:
                    cursor.execute("""
                        SELECT id FROM cy_enduser_master
                        WHERE enduser_code = %s AND id != %s
                    """, [enduser_code, enduser_id])
                    if cursor.fetchone():
                        messages.error(request, '같은 코드의 EndUser가 이미 존재합니다.')
                        return redirect('core:enduser_master_edit', enduser_id=enduser_id)
                
                cursor.execute("""
                    UPDATE cy_enduser_master
                    SET enduser_code = %s,
                        enduser_name = %s,
                        description = %s,
                        is_active = %s,
                        updated_at = NOW()
                    WHERE id = %s
                """, [enduser_code, enduser_name, description, is_active, enduser_id])
                
                messages.success(request, f'EndUser가 수정되었습니다: {enduser_code}')
                return redirect('core:enduser_master_list')
            except Exception as e:
                messages.error(request, f'오류: {str(e)}')
    
    context = {
        'enduser': enduser_data,
    }
    return render(request, 'core/enduser_master_edit.html', context)


@login_required
@user_passes_test(is_staff)
@require_POST
def enduser_master_delete(request, enduser_id):
    """EndUser 마스터 삭제"""
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT enduser_code FROM cy_enduser_master WHERE id = %s
        """, [enduser_id])
        result = cursor.fetchone()
        
        if result:
            enduser_code = result[0]
            cursor.execute("DELETE FROM cy_enduser_master WHERE id = %s", [enduser_id])
            messages.success(request, f'EndUser가 삭제되었습니다: {enduser_code}')
        else:
            messages.error(request, 'EndUser를 찾을 수 없습니다.')
    
    return redirect('core:enduser_master_list')
