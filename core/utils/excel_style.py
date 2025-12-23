# -*- coding: utf-8 -*-
"""
엑셀 출력 공통 스타일링 유틸리티
A4 용지 기준 인쇄에 최적화된 스타일 적용
"""
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins


def get_excel_styles():
    """공통 스타일 객체 반환"""
    thin_border = Border(
        left=Side(style='thin', color='B0B0B0'),
        right=Side(style='thin', color='B0B0B0'),
        top=Side(style='thin', color='B0B0B0'),
        bottom=Side(style='thin', color='B0B0B0')
    )
    
    header_font = Font(bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    data_font = Font(size=9)
    data_alignment = Alignment(vertical="center")
    data_alignment_center = Alignment(horizontal="center", vertical="center")
    data_alignment_right = Alignment(horizontal="right", vertical="center")
    
    # 짝수 행 배경 (얼룩말 무늬)
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    return {
        'thin_border': thin_border,
        'header_font': header_font,
        'header_fill': header_fill,
        'header_alignment': header_alignment,
        'data_font': data_font,
        'data_alignment': data_alignment,
        'data_alignment_center': data_alignment_center,
        'data_alignment_right': data_alignment_right,
        'alt_fill': alt_fill,
    }


def apply_header_style(ws, headers, row=1):
    """
    헤더 행에 스타일 적용 및 열 너비 설정
    
    Args:
        ws: 워크시트 객체
        headers: [(헤더명, 너비), ...] 또는 [헤더명, ...] 리스트
        row: 헤더 행 번호 (기본값 1)
    """
    styles = get_excel_styles()
    
    for col_idx, header_info in enumerate(headers, 1):
        if isinstance(header_info, tuple):
            header_name, width = header_info
        else:
            header_name = header_info
            width = 12  # 기본 너비
        
        cell = ws.cell(row=row, column=col_idx, value=header_name)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = styles['header_alignment']
        cell.border = styles['thin_border']
        
        # 열 너비 설정
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width
    
    # 헤더 행 높이
    ws.row_dimensions[row].height = 22


def apply_data_style(ws, start_row, end_row, num_cols, center_cols=None, right_cols=None):
    """
    데이터 행에 스타일 적용
    
    Args:
        ws: 워크시트 객체
        start_row: 데이터 시작 행
        end_row: 데이터 끝 행
        num_cols: 열 개수
        center_cols: 가운데 정렬할 열 인덱스 리스트 (1-based)
        right_cols: 오른쪽 정렬할 열 인덱스 리스트 (1-based)
    """
    styles = get_excel_styles()
    center_cols = center_cols or []
    right_cols = right_cols or []
    
    for row_idx in range(start_row, end_row + 1):
        is_alt = (row_idx - start_row) % 2 == 1
        ws.row_dimensions[row_idx].height = 18
        
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = styles['data_font']
            cell.border = styles['thin_border']
            
            # 정렬
            if col_idx in center_cols:
                cell.alignment = styles['data_alignment_center']
            elif col_idx in right_cols:
                cell.alignment = styles['data_alignment_right']
            else:
                cell.alignment = styles['data_alignment']
            
            # 얼룩말 무늬 배경
            if is_alt:
                cell.fill = styles['alt_fill']


def setup_print_area(ws, num_rows, num_cols, landscape=True, fit_to_page=True):
    """
    A4 인쇄 설정
    
    Args:
        ws: 워크시트 객체
        num_rows: 총 행 수
        num_cols: 총 열 수
        landscape: 가로 방향 여부 (기본값 True)
        fit_to_page: 용지 너비에 맞춤 여부 (기본값 True)
    """
    # 인쇄 영역 설정
    last_col = get_column_letter(num_cols)
    ws.print_area = f'A1:{last_col}{num_rows}'
    
    # 헤더 행 반복 인쇄
    ws.print_title_rows = '1:1'
    
    # 페이지 설정
    ws.page_setup.orientation = 'landscape' if landscape else 'portrait'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    
    if fit_to_page:
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0  # 높이는 자동
    
    # 여백 설정 (인치 단위, A4에 적합한 좁은 여백)
    ws.page_margins = PageMargins(
        left=0.4,    # 약 1cm
        right=0.4,
        top=0.5,
        bottom=0.5,
        header=0.2,
        footer=0.2
    )
    
    # 중앙 정렬 인쇄
    ws.print_options.horizontalCentered = True


def style_info_sheet(ws):
    """필터정보 시트 스타일 적용"""
    styles = get_excel_styles()
    
    # A열(라벨) 스타일
    for row in range(1, ws.max_row + 1):
        cell_a = ws.cell(row=row, column=1)
        cell_a.font = Font(bold=True, size=10)
        cell_a.alignment = Alignment(horizontal="right", vertical="center")
        
        cell_b = ws.cell(row=row, column=2)
        cell_b.font = Font(size=10)
        cell_b.alignment = Alignment(vertical="center")
    
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 50

